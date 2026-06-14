# -*- coding: utf-8 -*-
"""
news_research.py
================================
模块定位: 每日看盘报告 [资讯支柱] —— 为 ai final.py 的 Excel 看板增补两页:
    1) 📰 财经要闻 : 本周突发/头条 + 本月主题 (政策/产业/宏观/地缘/资源)
    2) 🏦 投行研报观点 : 高盛/大摩/小摩/瑞银/美银/巴克莱等的最新观点汇总

数据源与诚实声明 (重要)
---------------------------------------------------------------------------
  · 全部用 Google News RSS 实时抓取 (免 key / 支持中文任意关键词 / 带发布时间, 可按
    本周/本月过滤)。时效性 = 你运行脚本那一刻的最新资讯。
  · 投行【原始研报 PDF 是付费墙内的, 免费爬不到】。本页是【财经媒体对投行观点的公开转述】
    (如"高盛上调美光目标价"这类报道), 已在页头如实标注 —— 它是二手转述, 不是一手研报。
  · 源可能临时限流/改版; 抓不到时页面会显示提示而非崩溃 (各页 try 包裹)。

依赖: requests (ai.final 已用); xml 解析用标准库, 无新增重依赖。
"""

from __future__ import annotations

import time
import urllib.parse
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from typing import Dict, List, Optional

try:
    import requests
except Exception:
    requests = None

_HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}


# ==========================================================================
#   抓取底座: Google News RSS
# ==========================================================================
def _gnews(query: str, days: int, max_items: int = 8) -> List[dict]:
    """
    抓取某关键词的 Google News 中文 RSS, 过滤到最近 days 天, 返回
    [{title, source, dt, date, link}], 按时间倒序。失败返回 []。
    """
    if requests is None:
        return []
    url = ("https://news.google.com/rss/search?q="
           + urllib.parse.quote(query)
           + "&hl=zh-CN&gl=CN&ceid=CN:zh-Hans")
    try:
        r = requests.get(url, headers=_HEADERS, timeout=12)
        root = ET.fromstring(r.content)
    except Exception:
        return []

    now = datetime.now(timezone.utc)
    out: List[dict] = []
    for it in root.findall(".//item"):
        title = (it.findtext("title") or "").strip()
        link = (it.findtext("link") or "").strip()
        src_el = it.find("source")
        source = (src_el.text.strip() if src_el is not None and src_el.text else "")
        # Google News 标题常为 "实际标题 - 来源", 拆出来源兜底
        if not source and " - " in title:
            title, source = title.rsplit(" - ", 1)
        pub = it.findtext("pubDate")
        dt: Optional[datetime] = None
        try:
            dt = parsedate_to_datetime(pub) if pub else None
        except Exception:
            dt = None
        if dt is not None:
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            if (now - dt).days > days:
                continue
        out.append({
            "title": title.strip(), "source": source.strip(),
            "dt": dt, "date": dt.strftime("%m-%d") if dt else "—", "link": link,
        })
    out.sort(key=lambda x: x["dt"] or now, reverse=True)
    return out[:max_items]


def _collect(queries: List[str], days: int, max_items: int) -> List[dict]:
    """多关键词合并 + 按标题去重 + 时间倒序, 取前 max_items 条。"""
    seen, merged = set(), []
    for q in queries:
        for it in _gnews(q, days, max_items=8):
            key = it["title"][:24]
            if key in seen or not it["title"]:
                continue
            seen.add(key)
            merged.append(it)
        time.sleep(0.2)
    merged.sort(key=lambda x: x["dt"] or datetime.now(timezone.utc), reverse=True)
    return merged[:max_items]


# ==========================================================================
#   主题配置 (投资者爱关注的几条线)
# ==========================================================================
WEEK_QUERIES: List[str] = ["全球市场 突发", "A股 大盘", "美股 美联储", "地缘 冲突 市场"]

MONTH_THEMES: Dict[str, str] = {
    "政策 · 货币/财政": "货币政策 OR 美联储 OR 央行 降息 OR 财政政策",
    "产业 · 趋势风口":  "AI OR 半导体 OR 新能源 OR 算力 产业趋势",
    "宏观 · 经济环境":  "宏观经济 OR 通胀 OR GDP OR 就业 数据",
    "地缘 · 战争冲突":  "地缘 冲突 OR 战争 OR 制裁",
    "资源 · 大宗商品":  "原油 OR 黄金 OR 大宗商品 OR 稀土 OR 铜",
}

BANKS: Dict[str, str] = {
    "高盛 Goldman Sachs":   "高盛 (目标价 OR 评级 OR 研报 OR 展望 OR 预测)",
    "摩根士丹利 (大摩)":     "摩根士丹利 OR 大摩 (目标价 OR 评级 OR 展望)",
    "摩根大通 (小摩)":       "摩根大通 OR 小摩 (目标价 OR 评级 OR 展望)",
    "瑞银 UBS":             "瑞银 OR UBS (目标价 OR 评级 OR 展望)",
    "美银 BofA":            "美银 OR 美国银行 (目标价 OR 评级 OR 展望)",
    "巴克莱 Barclays":       "巴克莱 (目标价 OR 评级 OR 展望)",
    "中金 / 中信 (中资旗舰)": "中金公司 OR 中信证券 (目标价 OR 评级 OR 研报)",
}


# ==========================================================================
#   数据获取入口
# ==========================================================================
def fetch_news() -> dict:
    """返回 {'week': [...], 'themes': {主题: [...]}}。"""
    week = _collect(WEEK_QUERIES, days=7, max_items=12)
    themes = {label: _collect([q], days=31, max_items=6) for label, q in MONTH_THEMES.items()}
    return {"week": week, "themes": themes}


def fetch_research() -> Dict[str, List[dict]]:
    """返回 {投行: [...观点条目]}。"""
    return {bank: _collect([q], days=31, max_items=6) for bank, q in BANKS.items()}


# ==========================================================================
#   Excel 渲染 (报告式, 标题带超链接)
# ==========================================================================
def _styler():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    THIN = Side(style="thin", color="D9D9D9")
    return {
        "Font": Font, "Fill": PatternFill, "Align": Alignment,
        "BORDER": Border(left=THIN, right=THIN, top=THIN, bottom=THIN),
        "LEFT": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "CENTER": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }


def _news_rows(ws, r, items, S, empty_hint="（暂未取到，可能源临时限流，稍后重跑即可）"):
    """写一组新闻行: 日期 | 标题(超链接) | 来源。返回新的行号。"""
    if not items:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(row=r, column=1, value="   " + empty_hint)
        c.font = S["Font"](italic=True, size=9, color="999999")
        return r + 1
    for it in items:
        ws.cell(row=r, column=1, value=it["date"]).alignment = S["CENTER"]
        ws.cell(row=r, column=1).font = S["Font"](size=9, color="888888")
        tcell = ws.cell(row=r, column=2, value=it["title"])
        if it["link"]:
            tcell.hyperlink = it["link"]
            tcell.font = S["Font"](size=10, color="1155CC", underline="single")
        else:
            tcell.font = S["Font"](size=10)
        tcell.alignment = S["LEFT"]
        sc = ws.cell(row=r, column=3, value=it["source"])
        sc.font = S["Font"](size=9, color="666666"); sc.alignment = S["CENTER"]
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = S["BORDER"]
        r += 1
    return r


def _section_bar(ws, r, text, S, fill):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(row=r, column=1, value=text)
    c.fill = S["Fill"]("solid", fgColor=fill)
    c.font = S["Font"](bold=True, size=11, color="FFFFFF")
    c.alignment = S["LEFT"]
    ws.row_dimensions[r].height = 22
    return r + 1


def build_news_sheet(wb, data: Optional[dict] = None) -> None:
    """追加「📰 财经要闻」页: 本周焦点 + 本月五大主题。data 为空则内部实时抓取。"""
    S = _styler()
    if data is None:
        data = fetch_news()
    ws = wb.create_sheet("📰 财经要闻")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 82
    ws.column_dimensions["C"].width = 16

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = f"📰 财经要闻速览  ·  生成于 {datetime.now().strftime('%Y-%m-%d %H:%M')}  ·  来源: 全网新闻聚合(Google News)"
    t.font = S["Font"](bold=True, size=14, color="FFFFFF")
    t.fill = S["Fill"]("solid", fgColor="1F4E78"); t.alignment = S["CENTER"]
    ws.row_dimensions[1].height = 30
    r = 2

    r = _section_bar(ws, r, "🔥 本周焦点 / 突发头条 (近 7 天)", S, "C00000")
    r = _news_rows(ws, r, data.get("week", []), S)
    r += 1
    for label, items in data.get("themes", {}).items():
        r = _section_bar(ws, r, f"📌 本月 · {label} (近 31 天)", S, "375623")
        r = _news_rows(ws, r, items, S)
        r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    n = ws.cell(row=r, column=1,
                value="提示: 标题可点击跳原文。资讯为全网聚合, 仅供快速扫描, 重大决策请回到原文核实。")
    n.font = S["Font"](italic=True, size=9, color="999999")


def build_research_sheet(wb, data: Optional[Dict[str, List[dict]]] = None) -> None:
    """追加「🏦 投行研报观点」页: 各大投行最新观点。data 为空则内部实时抓取。"""
    S = _styler()
    if data is None:
        data = fetch_research()
    ws = wb.create_sheet("🏦 投行研报观点")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 82
    ws.column_dimensions["C"].width = 16

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = f"🏦 全球投行研报观点汇总  ·  生成于 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    t.font = S["Font"](bold=True, size=14, color="FFFFFF")
    t.fill = S["Fill"]("solid", fgColor="1F4E78"); t.alignment = S["CENTER"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    note = ws["A2"]
    note.value = ("⚠️ 诚实声明: 投行原始研报为付费墙内容, 此处为【财经媒体对其观点的公开转述】"
                  "(如目标价上调/评级变动的报道), 非一手 PDF。请据此判断, 重要决策回原文核实。")
    note.font = S["Font"](italic=True, size=9, color="C55A11"); note.alignment = S["LEFT"]
    ws.row_dimensions[2].height = 26
    r = 3

    for bank, items in data.items():
        r = _section_bar(ws, r, f"🏛️ {bank}", S, "1F4E78")
        r = _news_rows(ws, r, items, S, empty_hint="（近一月暂无公开转述，或源临时限流）")
        r += 1


# ==========================================================================
#   独立调试
# ==========================================================================
if __name__ == "__main__":
    print("抓取财经要闻 ...")
    nd = fetch_news()
    print(f"  本周焦点 {len(nd['week'])} 条")
    for it in nd["week"][:5]:
        print(f"   [{it['date']}] {it['title'][:50]} ({it['source']})")
    for label, items in nd["themes"].items():
        print(f"  {label}: {len(items)} 条")
    print("\n抓取投行观点 ...")
    rd = fetch_research()
    for bank, items in rd.items():
        print(f"  {bank}: {len(items)} 条")
        for it in items[:2]:
            print(f"   [{it['date']}] {it['title'][:50]} ({it['source']})")
