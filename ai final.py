# -*- coding: utf-8 -*-
# ==========================================================================
#   PETER 全资产量化决策看板 · 完全体 (ai final.py)
#   合并自：ai1.py（多流派智囊分析引擎） + ai2改改版.py（情绪爬虫 + 宏观晴雨表）
#   功能：1) 全球宏观晴雨表  2) 14 持仓量化扫描  3) 东财散户热议
#         4) openpyxl 金融机构级三标签页 Excel 自动落桌面并弹出
#   依赖：pip install yfinance ccxt pandas requests beautifulsoup4 openpyxl
# ==========================================================================
import os
import sys
import time
import shutil
import tempfile

import numpy as np
import pandas as pd

# ---------- 第三方库容错导入（缺啥都不让程序整体崩掉） ----------
try:
    import yfinance as yf
except Exception:
    yf = None
try:
    import ccxt
except Exception:
    ccxt = None
try:
    import requests
    from bs4 import BeautifulSoup
except Exception:
    requests = None
    BeautifulSoup = None

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ---------- 绘图库容错导入（缺失则自动降级为纯文字看板，绝不崩） ----------
try:
    import matplotlib
    matplotlib.use("Agg")  # 无界面后端，纯出图
    import matplotlib.pyplot as plt
    from matplotlib.transforms import blended_transform_factory
    import mplfinance as mpf
    # 中文字体（本机已确认含 Arial Unicode MS）
    matplotlib.rcParams["font.sans-serif"] = ["Arial Unicode MS", "Heiti TC", "STHeiti", "sans-serif"]
    matplotlib.rcParams["axes.unicode_minus"] = False
    _HAS_CHART = True
except Exception:
    _HAS_CHART = False

pd.set_option('display.unicode.east_asian_width', True)  # 终端中文对齐


# ==========================================================================
#   全局配置：14 个持仓 + 5 大宏观指数
# ==========================================================================
# 12 只个股 / ETF（港股统一最稳健的 4 位代码）
ASSETS = {
    "瑞芯微":        ("603893.SS", "A股"),
    "艾为电子":      ("688798.SS", "A股"),
    "中芯国际":      ("0981.HK",   "港股"),
    "科创芯片ETF":   ("588780.SS", "A股"),
    "纳指ETF广发":   ("159941.SZ", "A股"),
    "日经225ETF":    ("513880.SS", "A股"),
    "标普500ETF":    ("513500.SS", "A股"),
    "恒生科技ETF":   ("513130.SS", "A股"),
    "黄金ETF华安":   ("518880.SS", "A股"),
    "卫星ETF永赢":   ("159206.SZ", "A股"),
    "速腾聚创":      ("2498.HK",   "港股"),
    "腾讯控股":      ("0700.HK",   "港股"),
}
# 2 个币圈资产（欧易 OKX）→ 与上面 12 只合计正好 14 个持仓
CRYPTO = {
    "比特币 BTC/USDT": "BTC/USDT",
    "以太坊 ETH/USDT": "ETH/USDT",
}

# 中长线加仓引擎（midterm_position_engine）组合页的总弹药：
#   设为数字（如 50000）则 Excel 组合分配页显示建议金额；None = 只显示百分比。
MIDTERM_RESERVE = None
# A股四大核心指数 —— 创业板/科创50 因雅虎无指数历史序列，用对应 ETF 代理（走势≈一致且可买）
INDEX = {
    "上证综指":           ("000001.SS", "A股大盘"),
    "深证成指":           ("399001.SZ", "A股大盘"),
    "创业板指(ETF代理)":  ("159915.SZ", "创业板"),
    "科创50(ETF代理)":    ("588000.SS", "科创板"),
}

# 核心宏观指数（SPY / 费半 / A股四大指数 / 日经 / 恒生）
#   注：原代码把 000001.SS 误标为"富时中国A50"，其实它是上证综指，此处已并入 INDEX 一并纠正。
MACRO_INDEX = {
    "美股标普500 (SPY)":  ("^GSPC", "美股"),
    "费城半导体 (SOXX)":  ("^SOX",  "美股科技"),
    **INDEX,
    "日经225指数":        ("^N225", "日股"),
    "恒生指数":           ("^HSI",  "港股大盘"),
}


# ==========================================================================
#   模块 1：数据抓取底座（全程 try-except，绝不抛错）
# ==========================================================================
def normalize_code(code):
    """港股代码统一掰成雅虎认的 4 位格式，如 00981.HK -> 0981.HK"""
    if isinstance(code, str) and code.endswith(".HK"):
        num = code.replace(".HK", "").lstrip("0")
        return num.zfill(4) + ".HK"
    return code


def fetch_stock(code):
    """抓取股票 / ETF / 指数过去 100 天日 K 线，失败返回 None"""
    if yf is None:
        return None
    code = normalize_code(code)
    try:
        df = yf.download(code, period="100d", interval="1d",
                         auto_adjust=False, progress=False)
        if df is None or df.empty:
            return None
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        return df[["Open", "High", "Low", "Close", "Volume"]].dropna()
    except Exception as e:
        print(f"   ⚠️ 抓取 {code} 异常：{e}")
        return None


def fetch_crypto_okx(symbol):
    """免密抓取 100 天日 K 线：依次尝试 OKX→币安→Gate→Kraken，第一个成功即返回；全失败才 None。
    （OKX 公开行情常间歇性抽风导致 ccxt 报错，多所兜底保证 BTC/ETH 一定能进透视镜）"""
    if ccxt is None:
        return None
    for ex_id in ("okx", "binance", "gateio", "kraken"):
        try:
            ex = getattr(ccxt, ex_id)({'timeout': 10000, 'enableRateLimit': True})
            ohlcv = ex.fetch_ohlcv(symbol, timeframe="1d", limit=100)
            if not ohlcv:
                continue
            df = pd.DataFrame(ohlcv, columns=["Time", "Open", "High", "Low", "Close", "Volume"])
            df["Time"] = pd.to_datetime(df["Time"], unit="ms")
            return df.set_index("Time")
        except Exception:
            continue
    print(f"   ⚠️ 全部交易所抓取 {symbol} 均失败")
    return None


# ==========================================================================
#   模块 2：量化指标 + 改良版黄金买点信号
# ==========================================================================
def calc_indicators(df):
    """计算最新一天的 RSI(14) / MACD柱 / 布林带所属区间"""
    if df is None or len(df) < 20:
        return None
    try:
        close = df["Close"]

        # --- 14 日 RSI ---
        delta = close.diff()
        gain = delta.clip(lower=0).rolling(14).mean()
        loss = (-delta.clip(upper=0)).rolling(14).mean()
        rs = gain / (loss + 1e-9)
        rsi = 100 - 100 / (1 + rs)

        # --- MACD (12, 26, 9) ---
        ema12 = close.ewm(span=12, adjust=False).mean()
        ema26 = close.ewm(span=26, adjust=False).mean()
        dif = ema12 - ema26
        dea = dif.ewm(span=9, adjust=False).mean()
        macd_bar = (dif - dea) * 2

        # --- 布林带 (20, 2σ) ---
        ma20 = close.rolling(20).mean()
        std20 = close.rolling(20).std()
        upper = ma20 + 2 * std20
        lower = ma20 - 2 * std20

        last = close.iloc[-1]
        if last > upper.iloc[-1]:
            boll_pos = "超买预警 🔴(上轨上方)"
        elif last < lower.iloc[-1]:
            boll_pos = "超卖低吸 🟢(下轨下方)"
        else:
            boll_pos = "常态震荡区间 ⚪"

        return {
            "close": round(float(last), 3),
            "rsi": round(float(rsi.iloc[-1]), 1),
            "macd": round(float(macd_bar.iloc[-1]), 3),
            "boll_pos": boll_pos,
        }
    except Exception as e:
        print(f"   ⚠️ 指标计算异常：{e}")
        return None


def detect_golden_signal(df):
    """改良版黄金买点：连续 3 根阴线 + 第 4 根看涨放量吞没大阳线"""
    if df is None or len(df) < 5:
        return False
    try:
        o = df["Open"].values
        c = df["Close"].values
        v = df["Volume"].values
        three_red = all(c[i] < o[i] for i in [-4, -3, -2])
        bullish_engulf = (c[-1] > o[-1]) and (c[-1] > o[-2])
        vol_burst = v[-1] > (v[-4] + v[-3] + v[-2]) / 3
        return bool(three_red and bullish_engulf and vol_burst)
    except Exception:
        return False


# ==========================================================================
#   模块 2.5：ai1.py 原汁原味的多流派智囊引擎（一个结果都不省略）
# ==========================================================================
def calc_core_indicators(df):
    """算出 ai1 分析所需的全部核心指标"""
    c = df["Close"]
    # RSI(14)
    delta = c.diff()
    gain = delta.clip(lower=0).rolling(14).mean()
    loss = (-delta.clip(upper=0)).rolling(14).mean()
    rsi = float((100 - 100 / (1 + gain / (loss + 1e-9))).iloc[-1])
    # MACD
    dif = c.ewm(span=12, adjust=False).mean() - c.ewm(span=26, adjust=False).mean()
    dea = dif.ewm(span=9, adjust=False).mean()
    macd = float((dif.iloc[-1] - dea.iloc[-1]) * 2)
    # 布林带
    ma20 = c.rolling(20).mean()
    std20 = c.rolling(20).std()
    up = float((ma20 + 2 * std20).iloc[-1])
    low = float((ma20 - 2 * std20).iloc[-1])
    mid = float(ma20.iloc[-1])
    # 均线多空
    ma5 = float(c.rolling(5).mean().iloc[-1])
    ma60 = float(c.rolling(60).mean().iloc[-1]) if len(c) >= 60 else float(c.mean())
    # 从 100 天最高点的回撤幅度（给马丁策略用）
    drawdown = float((c.iloc[-1] / c.max() - 1) * 100)
    return {"price": float(c.iloc[-1]), "rsi": rsi, "macd": macd, "up": up,
            "low": low, "mid": mid, "ma5": ma5, "ma60": ma60, "dd": drawdown}


def detect_candle(df):
    """日本蜡烛图：识别最新形态，返回中文描述"""
    o, c, h, l = df["Open"].values, df["Close"].values, df["High"].values, df["Low"].values
    body = abs(c[-1] - o[-1])
    rng = h[-1] - l[-1] + 1e-9
    lower_shadow = min(o[-1], c[-1]) - l[-1]
    upper_shadow = h[-1] - max(o[-1], c[-1])
    if body / rng < 0.1:
        return "🌟十字星(多空胶着,变盘前兆)"
    if lower_shadow > body * 2 and upper_shadow < body:
        return "🔨锤子线(下方承接强,看涨反转)"
    if upper_shadow > body * 2 and lower_shadow < body:
        return "💫流星线(上方抛压重,看跌警告)"
    if c[-1] > o[-1] and c[-2] < o[-2] and c[-1] > o[-2] and o[-1] < c[-2]:
        return "🟢看涨吞没(多头反攻,强烈看涨)"
    if c[-1] < o[-1] and c[-2] > o[-2] and c[-1] < o[-2] and o[-1] > c[-2]:
        return "🔴看跌吞没(空头压制,警惕回调)"
    if all(c[i] < o[i] for i in [-3, -2, -1]):
        return "🐦‍⬛三只乌鸦(连续阴跌,趋势转弱)"
    return "➖普通K线(无明显形态)"


def analyze_asset(df):
    """🧠 融合四大流派(趋势/价值/蜡烛图/马丁)，返回结构化研判结果"""
    if df is None or len(df) < 26:
        return None
    ind = calc_core_indicators(df)
    candle = detect_candle(df)

    # ① 道氏趋势理论(均线+MACD)
    if ind["ma5"] > ind["ma60"] and ind["macd"] > 0:
        trend = "MA5上穿长均线且MACD红柱,多头排列,顺势可持有"
    elif ind["ma5"] < ind["ma60"] and ind["macd"] < 0:
        trend = "均线空头排列+MACD绿柱,下降趋势中,逆势抄底需谨慎"
    else:
        trend = "均线与MACD方向不一,处于震荡整理,观望为主"

    # ② 价值投资/逆向视角(RSI+布林)
    if ind["rsi"] > 70 or ind["price"] > ind["up"]:
        value = f"RSI={ind['rsi']:.0f}超买/突破布林上轨,估值偏贵,获利盘可减仓"
    elif ind["rsi"] < 30 or ind["price"] < ind["low"]:
        value = f"RSI={ind['rsi']:.0f}超卖/跌破布林下轨,情绪恐慌,逆向布局良机"
    else:
        value = f"RSI={ind['rsi']:.0f}处于中性区,价格在合理估值带内"

    # ③ 日本蜡烛图
    sakata = f"最新形态: {candle}"

    # ④ 马丁格尔加仓策略(基于回撤)
    dd = ind["dd"]
    if dd < -25:
        martin = f"已从高点回撤{dd:.1f}%,深度调整,可启动第3档分批补仓摊低成本"
    elif dd < -15:
        martin = f"回撤{dd:.1f}%,可启动第2档小幅补仓,严格控制仓位"
    elif dd < -8:
        martin = f"回撤{dd:.1f}%,试探性首档补仓,留足子弹"
    else:
        martin = f"距高点仅{dd:.1f}%,处高位区,不宜补仓,等回调"

    # 综合结论（与 ai1 完全一致：跨四条评论做多空关键词计数）
    comments = [f"【趋势派】{trend}", f"【价值派】{value}",
                f"【酒田战法】{sakata}", f"【马丁策略】{martin}"]
    bull = sum(1 for x in comments if any(k in x for k in ["看涨", "多头", "布局", "持有", "反攻"]))
    bear = sum(1 for x in comments if any(k in x for k in ["看跌", "空头", "减仓", "回调", "谨慎"]))
    if bull > bear:
        verdict = "🟢 偏多, 可逢低关注"
    elif bear > bull:
        verdict = "🔴 偏空, 注意风险控制"
    else:
        verdict = "🟡 中性, 震荡区间高抛低吸"

    return {**ind, "candle": candle, "trend": trend, "value": value,
            "sakata": sakata, "martin": martin, "verdict": verdict}


def multi_period_stats(df):
    """ai1 的多周期透视：1/7/30/100 天的区间涨跌、最高、最低、均量"""
    out = {}
    for key, n in [("d1", 1), ("d7", 7), ("d30", 30), ("d100", 100)]:
        seg = df.tail(n)
        chg = (seg["Close"].iloc[-1] / seg["Close"].iloc[0] - 1) * 100 if len(seg) > 1 else 0.0
        out[key] = {
            "chg": float(chg),
            "high": float(seg["High"].max()),
            "low": float(seg["Low"].min()),
            "vol": float(seg["Volume"].mean()),
        }
    return out


# ==========================================================================
#   模块 3：宏观晴雨表零件（VIX / 美债 / 机构风向标）
# ==========================================================================
def fetch_vix():
    """抓取 VIX 恐慌指数，返回 (数值, 中文状态串)"""
    if yf is not None:
        try:
            df = yf.download("^VIX", period="5d", interval="1d",
                             auto_adjust=False, progress=False)
            if df is not None and not df.empty:
                if isinstance(df.columns, pd.MultiIndex):
                    df.columns = df.columns.get_level_values(0)
                vix = float(df["Close"].iloc[-1])
                if vix > 30:
                    state = f"{round(vix, 2)} 🚨【极度恐慌！市场暴跌割肉，备好黄金/现金对冲】"
                elif vix > 20:
                    state = f"{round(vix, 2)} ⚠️【避险抬头！波动加大，分批网格，切勿满仓】"
                else:
                    state = f"{round(vix, 2)} ⚪【歌舞升平/平稳期！适合跟随趋势】"
                return vix, state
        except Exception:
            pass
    return 15.0, "15.0 ⚪【常态平稳区间（数据缺省值）】"


def fetch_global_switch():
    """抓取美债 10 年收益率（全球资产定价之锚），返回 (数值串, 风向串)"""
    if yf is not None:
        try:
            df = yf.download("^TNX", period="5d", interval="1d",
                             auto_adjust=False, progress=False)
            if df is not None and not df.empty:
                if isinstance(df.columns, pd.MultiIndex):
                    df.columns = df.columns.get_level_values(0)
                tnx = float(df["Close"].iloc[-1])
                state = "🚨 高利率压制科技股" if tnx > 4.5 else "🟢 利率回落利好芯片/币圈"
                return f"{round(tnx, 2)}%", state
        except Exception:
            pass
    return "暂无响应", "保持常态观望"


def region_institutional_view(df, vix_val):
    """结合 K 线与 VIX 生成华尔街机构视角研判结论"""
    try:
        close = df["Close"]
        ma20 = close.rolling(20).mean().iloc[-1]
        curr = close.iloc[-1]
        if vix_val > 25:
            return "🏦 外资主力无差别抛售流动性资产，多看少动，等左侧极端超卖区。"
        if curr > ma20:
            return "🏦 多头牢牢控盘，主力沿 20 日线顺势做多，持股待涨/回踩右侧加仓。"
        return "🏦 跌破 20 日生命线，机构转防守，建议马丁网格分批低吸，拉长补仓间距。"
    except Exception:
        return "🏦 多空交织，常态化震荡市。"


# ==========================================================================
#   模块 4：散户情绪爬取（东方财富股吧 + 币圈恐慌贪婪）
# ==========================================================================
def fetch_crypto_sentiment():
    """币圈全网恐慌与贪婪指数"""
    if requests is None:
        return "获取失败(缺 requests)"
    try:
        res = requests.get("https://api.alternative.me/fng/", timeout=5).json()
        value = res['data'][0]['value']
        cls = res['data'][0]['value_classification']
        return f"{value} ({cls})"
    except Exception:
        return "获取失败"


def fetch_eastmoney_guba_comments(code_num):
    """实时爬取东方财富股吧最新散户热门标题（取 3 条）"""
    if requests is None or BeautifulSoup is None:
        return ["爬虫库缺失，未能获取散户评论"]
    clean_code = ''.join(filter(str.isdigit, str(code_num)))
    url = f"https://guba.eastmoney.com/list,{clean_code}.html"
    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}
    try:
        res = requests.get(url, headers=headers, timeout=5)
        res.encoding = 'utf-8'  # 强行统一编码，消灭天书乱码
        soup = BeautifulSoup(res.text, 'html.parser')
        titles = [a.text.strip() for a in soup.select('.article_blk a') if a.text.strip()][:3]
        if not titles:
            titles = [a.text.strip() for a in soup.select('.title a') if a.text.strip()][:3]
        return titles if titles else ["暂无最新热门讨论"]
    except Exception:
        return ["连接股吧爬虫超时，未能获取到散户评论"]


# ==========================================================================
#   模块 5：openpyxl 金融机构级排版工具
# ==========================================================================
# 调色板
FILL_BLUE  = PatternFill("solid", fgColor="DDEBF7")   # 淡蓝表头（宏观）
FILL_GREEN = PatternFill("solid", fgColor="E2EFDA")   # 淡绿表头（持仓）
FILL_YELLOW_HDR = PatternFill("solid", fgColor="FFF2CC")  # 淡黄表头（情绪）
FILL_ALERT = PatternFill("solid", fgColor="FFFF00")   # 🔥 黄金买点整行警报：明亮黄
FILL_TITLE = PatternFill("solid", fgColor="1F4E78")   # 深蓝大标题底

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws, row, n_cols, fill):
    """把某一行设为表头：填充 + 加粗居中 + 边框"""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = Font(bold=True, size=12, color="333333")
        cell.alignment = CENTER
        cell.border = BORDER


def autofit_columns(ws, widths):
    """按给定宽度设置列宽"""
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def build_macro_sheet(ws, vix_state, tnx_str, tnx_view, macro_rows):
    """Sheet 1 · 宏观晴雨表"""
    ws.sheet_view.showGridLines = False
    autofit_columns(ws, [28, 22, 58])

    # 第 1 行：合并大标题（VIX 状态）
    ws.merge_cells("A1:C1")
    t1 = ws["A1"]
    t1.value = f"😱 今日 VIX 恐慌指数： {vix_state}"
    t1.font = Font(bold=True, size=15, color="FFFFFF")
    t1.fill = FILL_TITLE
    t1.alignment = CENTER
    ws.row_dimensions[1].height = 34

    # 第 2 行：合并（美债 10 年收益率）
    ws.merge_cells("A2:C2")
    t2 = ws["A2"]
    t2.value = f"💵 美债10年收益率（全球资产定价之锚）： {tnx_str}   |   🧭 {tnx_view}"
    t2.font = Font(bold=True, size=13, color="FFFFFF")
    t2.fill = FILL_TITLE
    t2.alignment = CENTER
    ws.row_dimensions[2].height = 28

    # 第 3 行：表头
    headers = ["核心指数名称", "最新点位", "机构风向标共识判定"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers), FILL_BLUE)

    # 第 4 行起：5 大指数数据
    r = 4
    for name, price, view in macro_rows:
        ws.cell(row=r, column=1, value=name).alignment = LEFT
        ws.cell(row=r, column=2, value=price).alignment = CENTER
        ws.cell(row=r, column=3, value=view).alignment = LEFT
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = BORDER
        ws.row_dimensions[r].height = 30
        r += 1


FILL_GROUP   = PatternFill("solid", fgColor="375623")   # 深绿 · 资产标题条
FILL_SECTION = PatternFill("solid", fgColor="A9D08E")   # 中绿 · 小节标题
FILL_TBLHDR  = PatternFill("solid", fgColor="EDEDED")   # 浅灰 · 日K表头
FILL_BULL    = PatternFill("solid", fgColor="C6EFCE")   # 综合研判·偏多绿
FILL_BEAR    = PatternFill("solid", fgColor="FFC7CE")   # 综合研判·偏空红
FILL_NEU     = PatternFill("solid", fgColor="FFEB9C")   # 综合研判·中性黄


def _fmt(v, kind):
    """数字格式化小助手"""
    if v is None:
        return "—"
    if kind == "price":
        return f"{v:,.3f}"
    if kind == "rsi":
        return f"{v:.1f}"
    if kind == "macd":
        return f"{v:.3f}"
    if kind == "pct":
        return f"{v:+.2f}%"
    if kind == "dd":
        return f"{v:+.1f}%"
    if kind == "vol":
        return f"{v:,.0f}"
    return str(v)


def _merged_row(ws, r, text, fill=None, font=None, align=None, height=None, border=False):
    """整行合并 A:F，写入一段文字（报告式排版的基本砖块）"""
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cell = ws.cell(row=r, column=1, value=text)
    cell.fill = fill or PatternFill()
    cell.font = font or Font(size=11)
    cell.alignment = align or LEFT
    if border:
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER
    if height:
        ws.row_dimensions[r].height = height
    return cell


def _adjust_splits(data):
    """把基金份额折算 / 股票拆分造成的“假断崖”做后复权，让日K连续。
    用『跳空』判定（当日整条K线都远离昨收）来识别折算，避免误伤普通大阴线与7×24的币。"""
    o, h, l, c = (data["Open"].values, data["High"].values,
                  data["Low"].values, data["Close"].values)
    f = np.ones(len(c))
    cum = 1.0
    for i in range(len(c) - 1, 0, -1):
        prev = c[i - 1]
        if prev > 0 and (h[i] < prev * 0.8 or l[i] > prev * 1.25):  # 整日跳空 → 视为折算/拆分
            cum *= c[i] / prev
        f[i - 1] = cum
    if np.allclose(f, 1.0):
        return data
    adj = data.copy()
    for col in ["Open", "High", "Low", "Close"]:
        adj[col] = adj[col].values * f
    return adj


def make_kline_chart(name, df, out_dir, idx):
    """生成日K蜡烛图：标准均线 + 成交量子图 + 支撑线 + 分批加仓点位，返回 PNG 路径"""
    if not _HAS_CHART or df is None or len(df) < 20:
        return None
    try:
        data = df.tail(100).copy()
        data.index = pd.to_datetime(data.index)
        data = data[["Open", "High", "Low", "Close", "Volume"]].astype(float)
        data = _adjust_splits(data)  # 🔧 后复权，干掉折算假断崖（专治科创芯片ETF这类）

        n = len(data)
        mavs = tuple(m for m in (5, 10, 20, 60) if m < n)  # 标准均线自适应

        lo_all = float(data["Low"].min())
        hi_all = float(data["High"].max())
        hi100 = float(data["High"].max())   # 100 日高点 → 马丁回撤基准

        # 关键位：近20日支撑 / 近60日支撑 / 马丁 8·15·25% 分批加仓位
        candidates = [
            (float(data["Low"].tail(20).min()),               "近20日支撑",   "#2E86C1", "-"),
            (float(data["Low"].tail(min(60, n)).min()),       "近60日支撑",   "#117A65", "-"),
            (hi100 * 0.92,                                     "首档加仓 -8%",  "#E67E22", "--"),
            (hi100 * 0.85,                                     "二档加仓 -15%", "#CA6F1E", "--"),
            (hi100 * 0.75,                                     "三档加仓 -25%", "#C0392B", "--"),
        ]
        # 只保留落在合理视野内的关键位（避免把蜡烛压扁）
        band_lo, band_hi = lo_all * 0.82, hi_all * 1.05
        kept = [(v, lab, col, ls) for (v, lab, col, ls) in candidates if band_lo <= v <= band_hi]

        hlines = dict(
            hlines=[v for v, *_ in kept],
            colors=[c for _, _, c, _ in kept],
            linestyle=[s for *_, s in kept],
            linewidths=[1.1] * len(kept),
        )

        style = mpf.make_mpf_style(
            base_mpf_style="yahoo",
            rc={"font.sans-serif": ["Arial Unicode MS", "Heiti TC", "STHeiti", "sans-serif"],
                "axes.unicode_minus": False},
        )

        fig, axes = mpf.plot(
            data, type="candle", volume=True, mav=mavs, style=style,
            figsize=(7.4, 4.3), returnfig=True, hlines=hlines,
            datetime_format="%m-%d", xrotation=0, tight_layout=True,
            ylabel="价格", ylabel_lower="成交量",
        )
        price_ax = axes[0]
        # 让所有关键位都进视野
        ys = [v for v, *_ in kept] + [lo_all, hi_all]
        price_ax.set_ylim(min(ys) * 0.985, max(ys) * 1.015)

        # 右侧贴标签（x 用坐标系比例，y 用真实价格）
        trans = blended_transform_factory(price_ax.transAxes, price_ax.transData)
        for v, lab, col, _ls in kept:
            price_ax.text(1.005, v, f"{lab} {v:.2f}", transform=trans, color=col,
                          fontsize=8, va="center", ha="left",
                          bbox=dict(boxstyle="round,pad=0.15", fc="white", ec=col, alpha=0.85))

        ma_txt = " / ".join("MA" + str(m) for m in mavs) if mavs else "均线"
        price_ax.set_title(f"{name} · 日K线（{ma_txt}）", fontsize=11, fontweight="bold")

        path = os.path.join(out_dir, f"k_{idx}.png")
        fig.savefig(path, dpi=110, bbox_inches="tight")
        plt.close(fig)
        return path
    except Exception as e:
        print(f"   ⚠️ {name} 日K图生成失败：{e}")
        try:
            plt.close("all")
        except Exception:
            pass
        return None


def build_holdings_sheet(ws, rows, chart_dir=None):
    """Sheet 2 · 持仓透视镜 —— 报告式纵向排版，原汁原味复刻 ai1.py 每个资产的全部输出。"""
    ws.sheet_view.showGridLines = False
    autofit_columns(ws, [15, 11, 11, 11, 11, 17])  # 日期 / 开 / 高 / 低 / 收 / 量

    WHITE_BOLD = Font(bold=True, size=13, color="FFFFFF")
    SEC_FONT   = Font(bold=True, size=11, color="1F3D14")
    LBL_FONT   = Font(bold=True, size=11, color="333333")
    TXT_FONT   = Font(size=11, color="222222")

    ws.column_dimensions["G"].width = 2.5  # A:F 报告 与 右侧K线图 之间留个小缝

    r = 1
    for idx, row in enumerate(rows):
        name = row["name"]
        ana = row["ana"]
        mp = row["mp"]
        df = row["df"]
        golden = row["golden"]
        start_r = r  # 记下本资产区块起始行，K线图就钉在这行的右侧

        # ===== ⓪ 右侧日K蜡烛图（浮动锚定 H 列，不占用左侧报告单元格）=====
        if chart_dir is not None:
            img_path = make_kline_chart(name, df, chart_dir, idx)
            if img_path:
                try:
                    pic = XLImage(img_path)
                    pic.width, pic.height = 660, 384
                    ws.add_image(pic, f"H{start_r}")
                except Exception as e:
                    print(f"   ⚠️ {name} K线图嵌入失败：{e}")

        # ===== ① 资产标题条 =====
        if golden:
            title = f"🔥💥 【{name}】 黄金放量大阳线买点触发！！  现价: {ana['price']:.3f} 💥🔥" if ana \
                else f"🔥💥 【{name}】 黄金买点触发！！💥🔥"
            _merged_row(ws, r, title, fill=FILL_ALERT,
                        font=Font(bold=True, size=13, color="C00000"), align=CENTER, height=30)
        else:
            price_txt = f"   现价: {ana['price']:.3f}" if ana else ""
            _merged_row(ws, r, f"📊 【{name}】 智囊分析报告{price_txt}",
                        fill=FILL_GROUP, font=WHITE_BOLD, align=CENTER, height=30)
        r += 1

        # 数据不足兜底
        if ana is None or mp is None or df is None or df.empty:
            _merged_row(ws, r, "   ⚠️ 数据不足或抓取失败，无法生成完整研判（需至少 26 根日K线）",
                        font=Font(italic=True, color="C00000"), height=22, border=True)
            r += 2  # 留一行空白
            continue

        # ===== ② 多周期透视 =====
        _merged_row(ws, r, "📅 多周期透视  (最近 1 / 7 / 30 / 100 天)",
                    fill=FILL_SECTION, font=SEC_FONT, align=LEFT, height=22, border=True)
        r += 1

        # 最近1天
        d1 = mp["d1"]
        _merged_row(ws, r, f"   [最近1天]   区间涨跌: {d1['chg']:+.2f}%    最新收盘: {ana['price']:.3f}",
                    font=LBL_FONT, height=20)
        r += 1

        # 最近7天 + 每日K线明细表（完整复刻 ai1 的逐根K线打印）
        d7 = mp["d7"]
        _merged_row(ws, r, f"   [最近7天]   区间涨跌: {d7['chg']:+.2f}%    最新收盘: {ana['price']:.3f}",
                    font=LBL_FONT, height=20)
        r += 1
        # 表头
        for col, h in enumerate(["日期", "开盘", "最高", "最低", "收盘", "成交量"], 1):
            c = ws.cell(row=r, column=col, value=h)
            c.fill = FILL_TBLHDR
            c.font = Font(bold=True, size=10)
            c.alignment = CENTER
            c.border = BORDER
        r += 1
        seg = df.tail(7)
        for idx, krow in seg.iterrows():
            try:
                date_str = idx.strftime("%Y-%m-%d")
            except Exception:
                date_str = str(idx)[:10]
            kvals = [date_str, f"{krow['Open']:.3f}", f"{krow['High']:.3f}",
                     f"{krow['Low']:.3f}", f"{krow['Close']:.3f}", f"{int(krow['Volume']):,}"]
            for col, v in enumerate(kvals, 1):
                c = ws.cell(row=r, column=col, value=v)
                c.font = Font(size=10)
                c.alignment = LEFT if col == 1 else CENTER
                c.border = BORDER
            r += 1

        # 最近30天 / 100天：区间 + 高低均量
        for label, key in [("最近30天", "d30"), ("最近100天", "d100")]:
            d = mp[key]
            _merged_row(ws, r,
                        f"   [{label}]  区间涨跌: {d['chg']:+.2f}%   最新收盘: {ana['price']:.3f}   "
                        f"高: {d['high']:.3f}   低: {d['low']:.3f}   均量: {d['vol']:,.0f}",
                        font=LBL_FONT, height=20)
            r += 1

        # ===== ③ 多流派智囊深度分析 =====
        _merged_row(ws, r, "🧠 多流派智囊深度分析  (趋势 · 价值 · 酒田 · 马丁)",
                    fill=FILL_SECTION, font=SEC_FONT, align=LEFT, height=22, border=True)
        r += 1
        bullets = [
            f"   ▪ 【趋势派】{ana['trend']}",
            f"   ▪ 【价值派】{ana['value']}",
            f"   ▪ 【酒田战法】{ana['sakata']}",
            f"   ▪ 【马丁策略】{ana['martin']}",
        ]
        for b in bullets:
            _merged_row(ws, r, b, font=TXT_FONT, align=LEFT, height=20)
            r += 1

        # ===== ④ 综合研判（按多空着色） =====
        verdict = ana["verdict"]
        if "偏多" in verdict:
            vfill = FILL_BULL
        elif "偏空" in verdict:
            vfill = FILL_BEAR
        else:
            vfill = FILL_NEU
        _merged_row(ws, r, f"   ● 综合研判: {verdict}",
                    fill=vfill, font=Font(bold=True, size=12, color="222222"),
                    align=LEFT, height=24, border=True)
        r += 1

        # 资产之间留一行空白做间隔
        ws.row_dimensions[r].height = 8
        r += 1


def build_sentiment_sheet(ws, rows):
    """Sheet 3 · 散户热议看点"""
    ws.sheet_view.showGridLines = False
    autofit_columns(ws, [22, 40, 40, 40])

    headers = ["资产名称", "东财热议 ①", "东财热议 ②", "东财热议 ③"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers), FILL_YELLOW_HDR)

    r = 2
    for name, titles in rows:
        ws.cell(row=r, column=1, value=name).alignment = LEFT
        for i in range(3):
            txt = titles[i] if i < len(titles) else ""
            ws.cell(row=r, column=2 + i, value=txt).alignment = LEFT
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER
        ws.row_dimensions[r].height = 30
        r += 1


# ==========================================================================
#   主引擎
# ==========================================================================
def main():
    print("🚀 启动 Peter 全资产量化决策看板 · 完全体 ...\n")

    # ---------- 阶段 A：全球宏观晴雨表 ----------
    print("=" * 70)
    print("🌍 阶段一 · 全球宏观晴雨表")
    print("=" * 70)

    vix_val, vix_state = fetch_vix()
    tnx_str, tnx_view = fetch_global_switch()
    print(f"😱 VIX 恐慌指数：{vix_state}")
    print(f"💵 美债10年收益率：{tnx_str}   🧭 {tnx_view}")

    macro_rows = []
    print("\n📊 五大核心市场 · 机构共识研判：")
    for name, (code, _mtype) in MACRO_INDEX.items():
        df = fetch_stock(code)
        if df is not None and not df.empty:
            price = round(float(df["Close"].iloc[-1]), 2)
            view = region_institutional_view(df, vix_val)
        else:
            price = "数据失败"
            view = "🏦 指数数据抓取失败，保持观望。"
        macro_rows.append((name, price, view))
        print(f"🔹 {name} | 最新点位: {price}\n   {view}")
        time.sleep(0.6)

    # ---------- 阶段 B：14 持仓量化扫描 + 散户情绪 ----------
    print("\n" + "=" * 70)
    print("📈 阶段二 · 14 持仓量化透视 + 散户热议")
    print("=" * 70)
    print(f"🪙 币圈恐慌与贪婪大盘：{fetch_crypto_sentiment()}")

    holdings_rows = []    # 给 Sheet 2
    sentiment_rows = []   # 给 Sheet 3

    # B1：先扫描 12 只个股 / ETF（带东财股吧爬取）—— 跑 ai1 全套智囊分析
    for name, (code, mtype) in ASSETS.items():
        print(f"\n🔍 {name} ({code}) - {mtype}")
        df = fetch_stock(code)
        ana = analyze_asset(df)
        mp = multi_period_stats(df) if (df is not None and not df.empty) else None
        golden = detect_golden_signal(df)

        if ana:
            print(f"   📈 现价: {ana['price']:.3f} | RSI: {ana['rsi']:.1f} | "
                  f"MACD: {ana['macd']:.3f} | {ana['candle']}")
            print(f"   🧠 {ana['verdict']}")
            if golden:
                print("   🔥💥 触发【改良版黄金放量大阳线买点】！！💥🔥")
        else:
            print("   ❌ 数据抓取失败或数据不足")
        holdings_rows.append({"name": name, "ana": ana, "mp": mp, "df": df, "golden": golden})

        titles = fetch_eastmoney_guba_comments(code)
        print("   💬 东财热议：" + " / ".join(titles))
        sentiment_rows.append((name, titles))
        time.sleep(1.0)

    # B2：再扫描 2 个币圈资产（欧易 OKX）
    crypto_sent = fetch_crypto_sentiment()
    for name, symbol in CRYPTO.items():
        print(f"\n🔍 {name} (OKX)")
        df = fetch_crypto_okx(symbol)
        ana = analyze_asset(df)
        mp = multi_period_stats(df) if (df is not None and not df.empty) else None
        golden = detect_golden_signal(df)

        if ana:
            print(f"   📈 欧易实时价: {ana['price']:.3f} | RSI: {ana['rsi']:.1f} | "
                  f"MACD: {ana['macd']:.3f} | {ana['candle']}")
            print(f"   🧠 {ana['verdict']}")
            if golden:
                print("   🔥💥 触发【改良版黄金放量大阳线买点】！！💥🔥")
        else:
            print("   ❌ 欧易 API 连接超时或数据不足")
        holdings_rows.append({"name": name, "ana": ana, "mp": mp, "df": df, "golden": golden})
        # 币圈无东财股吧，用恐慌贪婪指数兜底
        sentiment_rows.append((name, [
            f"币圈无东财股吧，参考全网恐慌贪婪指数：{crypto_sent}",
            "（数据源：alternative.me Fear & Greed Index）",
            "",
        ]))
        time.sleep(1.0)

    # B3：再扫描 4 大 A 股指数（大盘风向标）—— 同样跑四流派智囊 + K线，但不属个股持仓
    for name, (code, mtype) in INDEX.items():
        print(f"\n🔍 {name} ({code}) - {mtype} [指数]")
        df = fetch_stock(code)
        ana = analyze_asset(df)
        mp = multi_period_stats(df) if (df is not None and not df.empty) else None
        golden = detect_golden_signal(df)
        if ana:
            print(f"   📈 现价: {ana['price']:.3f} | RSI: {ana['rsi']:.1f} | "
                  f"MACD: {ana['macd']:.3f} | {ana['candle']}")
            print(f"   🧠 {ana['verdict']}")
        else:
            print("   ❌ 指数数据抓取失败或数据不足")
        holdings_rows.append({"name": name, "ana": ana, "mp": mp, "df": df, "golden": golden})
        sentiment_rows.append((name, ["大盘指数，无个股股吧（情绪请看宏观晴雨表 VIX/恐慌贪婪）", "", ""]))
        time.sleep(1.0)

    # ---------- 阶段 C：写入桌面 Excel ----------
    print("\n" + "=" * 70)
    print("📒 阶段三 · 正在生成金融机构级 Excel 决策看板 ...")
    print("=" * 70)

    save_path = os.path.expanduser("~/Desktop/Peter_今日量化决策看板.xlsx")
    chart_dir = None
    if _HAS_CHART:
        chart_dir = tempfile.mkdtemp(prefix="kline_")
        print("   🖼️ 正在绘制每个标的的日K蜡烛图（均线+成交量+支撑线+分批加仓位）...")
    else:
        print("   ⚠️ 未安装 matplotlib/mplfinance，本次降级为纯文字看板"
              "（如需K线图请：pip install mplfinance）")
    try:
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "🌍 宏观晴雨表"
        build_macro_sheet(ws1, vix_state, tnx_str, tnx_view, macro_rows)

        ws2 = wb.create_sheet("📈 持仓透视镜")
        build_holdings_sheet(ws2, holdings_rows, chart_dir=chart_dir)

        ws3 = wb.create_sheet("💬 散户热议看点")
        build_sentiment_sheet(ws3, sentiment_rows)

        # ---------- 阶段四：中长线加仓决策（融合 midterm_position_engine，完整不省略） ----------
        _mte_dec = []
        try:
            import midterm_position_engine as _mte
            print("\n" + "=" * 70)
            print("🧭 阶段四 · 中长线加仓决策引擎（半年~一年视角 · 多年日K · 统计口径）")
            print("=" * 70)
            _specs = ([(n, c, m, False) for n, (c, m) in ASSETS.items()]
                      + [(n, c, m, False) for n, (c, m) in INDEX.items()]
                      + [(n, s, "币圈", True) for n, s in CRYPTO.items()])
            _mte_dec = _mte.collect_decisions(_specs, verbose=True)
            _mte.build_midterm_sheets(wb, _mte_dec, reserve_capital=MIDTERM_RESERVE)
            print(f"   ✅ 已并入 {len(_mte_dec)} 个资产的完整中长线加仓决策页 + 组合弹药分配页")
        except Exception as _e:
            print(f"   ⚠️ 中长线加仓页生成失败（不影响其余看板）：{_e}")

        # ---------- 阶段五：基本面分析 + 三支柱综合研判（币圈不做基本面） ----------
        try:
            import fundamental_analysis as _fa
            print("\n" + "=" * 70)
            print("📊 阶段五 · 中长线基本面分析（四维因子库）+ 技术×统计×基本面 综合研判")
            print("=" * 70)
            # 基本面只对股票/ETF（"ETF" 在名字里即视为指数型，跳过个股财报）；4 大指数一律按指数型处理
            _fa_specs = ([(n, c, m, ("ETF" in n)) for n, (c, m) in ASSETS.items()]
                         + [(n, c, m, True) for n, (c, m) in INDEX.items()])
            _fa_reps = _fa.analyze_fundamentals(_fa_specs, verbose=True)
            _fa.build_fundamental_sheet(wb, _fa_reps)
            # 技术支柱：复用 ai1 智囊四流派研判（含股票与币圈）
            _tech_map = {row["name"]: row["ana"] for row in holdings_rows if row.get("ana")}
            _fa.build_summary_sheet(wb, _tech_map, _mte_dec, _fa_reps)
            print(f"   ✅ 已并入 {len(_fa_reps)} 个标的的基本面页 + 综合研判页")
        except Exception as _e:
            print(f"   ⚠️ 基本面/综合研判页生成失败（不影响其余看板）：{_e}")

        # ---------- 阶段六：财经要闻 + 投行研报观点（实时全网聚合） ----------
        try:
            import news_research as _nr
            print("\n" + "=" * 70)
            print("📰 阶段六 · 财经要闻（本周头条+本月主题）+ 全球投行研报观点")
            print("=" * 70)
            _nr.build_news_sheet(wb)
            print("   ✅ 已并入「📰 财经要闻」页")
            _nr.build_research_sheet(wb)
            print("   ✅ 已并入「🏦 投行研报观点」页")
        except Exception as _e:
            print(f"   ⚠️ 资讯/研报页生成失败（不影响其余看板）：{_e}")

        wb.save(save_path)
        print(f"✅ Excel 已生成：{save_path}")

        # 啪的一下弹出 Excel
        try:
            if sys.platform == "darwin":
                os.system(f'open "{save_path}"')
            elif sys.platform.startswith("win"):
                os.startfile(save_path)  # noqa
            else:
                os.system(f'xdg-open "{save_path}"')
        except Exception as e:
            print(f"   ⚠️ 自动打开失败（请手动打开桌面文件）：{e}")
    except Exception as e:
        print(f"❌ Excel 写入失败：{e}")
    finally:
        # 图片已嵌进 xlsx，临时 PNG 可以安全清理
        if chart_dir:
            shutil.rmtree(chart_dir, ignore_errors=True)

    print("\n✨ 全流程跑通！看大盘定方向，看个股定买卖。决策仅供参考！")


if __name__ == "__main__":
    main()
