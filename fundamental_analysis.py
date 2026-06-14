# -*- coding: utf-8 -*-
"""
fundamental_analysis.py
================================
模块定位: 中长线复利机器 [基本面支柱] —— 把财报数据提取成"华尔街给科技股定价的因子",
          打成一个可解释的基本面分数 + 关键指标专栏, 并产出"技术×统计×基本面"的最终综合研判。

设计取向 (诚实第一)
---------------------------------------------------------------------------
  · 币圈不做基本面 (BTC/ETH 没有营收/PE, 此模块只处理股票/ETF)。
  · 华尔街对【科技成长股】与【传统周期股】的定价逻辑不同 -> 本模块按行业(sector)选择不同
    的因子权重与阈值: 科技看 增长持续性 / 现金流转换 / 业绩确定性; 周期看 估值 / ROE / 股息。
  · 因子库分四维 (用户给定): ①增长 ②盈利与真实现金流 ③运营效率与粘性 ④前瞻指引。
    其中【免费数据源(yfinance/东财同花顺)能算的】真算; 【披露项(Billings/RPO/ARR/NRR/
    Guidance)免费源拿不到的】如实标注 "披露项·可手填", 由 MANUAL_OVERRIDES 选填, 绝不编造。
  · 护城河(moat): 由所处行业定性分档 + 用毛利率/ROE/研发强度做量化加强。
  · 政策与叙事: 用"消息面"(最近约一个月新闻标题) 代替。
  · 最终综合研判【三支柱并列, 不糊成一个数字】(三者常打架, 分开看才诚实), 同时给一个加权
    综合分仅作排序参考。

数据源:
    yfinance (主干, A股/港股/美股财报与估值均可) ; akshare (可选: 装了则增强 A股, 没装自动跳过)
依赖:
    pip install yfinance        # 必需
    pip install akshare         # 可选 (解锁 A股估值历史分位等, 缺失不影响主体)
"""

from __future__ import annotations

import time
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

try:
    import yfinance as yf
except Exception:
    yf = None
try:
    import akshare as ak   # 可选增强, 缺失自动降级
except Exception:
    ak = None


# ==========================================================================
#   手填配置: 免费源拿不到的"披露项"与定性补充 (想填就填, 不填则显示"可手填")
#   key 用 ai.final 的展示名; 值是字段->字符串。例:
#       MANUAL_OVERRIDES = {"腾讯控股": {"NRR": "—", "叙事": "AI+视频号商业化", "定性分": "4"}}
# ==========================================================================
MANUAL_OVERRIDES: Dict[str, Dict[str, str]] = {}


# ==========================================================================
#   行业画像: 把 yfinance 的 sector/industry 归到自有类别, 决定因子权重/阈值/护城河
# ==========================================================================
@dataclass(frozen=True)
class SectorProfile:
    key: str                        # 自有类别中文名
    gross_margin_target: float      # 该类"健康"毛利率基准 (打分用)
    moat_tier: str                  # 护城河默认档 (再用量化指标微调)
    moat_reason: str
    # 因子权重 (和为 1): 增长 / 毛利 / FCF / Rule40 / ROE / 估值 / 健康度
    w_growth: float; w_gm: float; w_fcf: float; w_rule40: float
    w_roe: float; w_val: float; w_health: float
    is_growth: bool                 # True=科技成长(容忍高估值/烧钱换增长); False=价值周期


_TECH_SEMI = SectorProfile("半导体设计/制造", 0.45, "中高",
                           "芯片设计的IP与流片壁垒 / 代工的资本与制程壁垒",
                           0.25, 0.15, 0.13, 0.15, 0.10, 0.12, 0.10, True)
_TECH_SW = SectorProfile("软件/互联网平台", 0.70, "高",
                         "网络效应 + 生态/数据壁垒 + 高切换成本",
                         0.22, 0.15, 0.18, 0.15, 0.10, 0.13, 0.07, True)
_TECH_HW = SectorProfile("科技硬件/智能制造", 0.30, "中低",
                         "技术领先窗口短, 易被迭代/价格战, 壁垒偏弱",
                         0.28, 0.12, 0.12, 0.13, 0.10, 0.15, 0.10, True)
_CYCLICAL = SectorProfile("周期/金融/能源", 0.30, "中",
                          "强周期, 壁垒多来自牌照/规模/资源禀赋",
                          0.12, 0.10, 0.12, 0.08, 0.20, 0.28, 0.10, False)
_DEFAULT = SectorProfile("通用", 0.35, "中",
                         "按通用价值框架评估",
                         0.18, 0.13, 0.15, 0.12, 0.14, 0.18, 0.10, False)


def _classify_sector(sector: str, industry: str) -> SectorProfile:
    s = (sector or "").lower()
    ind = (industry or "").lower()
    if "semiconductor" in s or "semiconductor" in ind:
        return _TECH_SEMI
    if "communication" in s or "software" in ind or "internet" in ind:
        return _TECH_SW
    if "technology" in s or "electronic" in ind or "hardware" in ind or "auto parts" in ind:
        return _TECH_HW
    if any(k in s for k in ["financial", "energy", "basic materials", "utilities", "industrials"]):
        return _CYCLICAL
    return _DEFAULT


# ==========================================================================
#   结果载体
# ==========================================================================
@dataclass(frozen=True)
class FundamentalReport:
    name: str
    market: str
    is_etf: bool
    ok: bool                            # 是否成功取得可用财报
    sector_cn: str
    sector_raw: str
    industry_raw: str
    score: float                        # 0~100 基本面综合分
    grade: str                          # A优 / B良 / C中 / D弱
    subscores: Dict[str, float]         # 各子项得分 (透明)
    metrics: List[Tuple[str, str, str]] # 关键指标专栏: (指标名, 数值, 评判)
    moat_tier: str
    moat_reason: str
    rule40: Optional[float]
    rule40_pass: Optional[bool]
    news: List[str]                     # 消息面: 最近约一个月新闻标题
    flags: List[str]                    # 数据缺失/排雷提示
    note: str = ""


# ==========================================================================
#   财报字段提取 (全程容错: 拿不到就 None, 绝不抛错)
# ==========================================================================
def _get_row(df, names: List[str]) -> Optional[float]:
    """从 yfinance 财报 DataFrame 取最近一期某行 (多别名兜底)。"""
    if df is None or getattr(df, "empty", True):
        return None
    for nm in names:
        if nm in df.index:
            try:
                val = df.loc[nm].dropna()
                if len(val):
                    return float(val.iloc[0])
            except Exception:
                continue
    return None


def _get_row_prev(df, names: List[str]) -> Optional[float]:
    """取上一期 (用于同比/环比变化, 如 Magic Number 的毛利增量)。"""
    if df is None or getattr(df, "empty", True):
        return None
    for nm in names:
        if nm in df.index:
            try:
                val = df.loc[nm].dropna()
                if len(val) >= 2:
                    return float(val.iloc[1])
            except Exception:
                continue
    return None


def _fetch_news(t) -> List[str]:
    """最近约一个月新闻标题 (代替政策/叙事面)。A股常无英文流, 缺失返回空。"""
    out: List[str] = []
    try:
        news = getattr(t, "news", None) or []
        cutoff = time.time() - 31 * 86400
        for item in news:
            # 兼容 yfinance 新旧两种结构
            content = item.get("content", item)
            title = content.get("title") or item.get("title")
            ts = item.get("providerPublishTime")
            if ts and ts < cutoff:
                continue
            if title:
                out.append(title.strip())
            if len(out) >= 5:
                break
    except Exception:
        pass
    return out


# ==========================================================================
#   打分小工具 (分段线性映射到 0~100, 全部可解释)
# ==========================================================================
def _piecewise(x: Optional[float], pts: List[Tuple[float, float]]) -> Optional[float]:
    """分段线性插值。pts=[(x0,score0),...] 升序。x 为 None -> None。"""
    if x is None:
        return None
    if x <= pts[0][0]:
        return pts[0][1]
    if x >= pts[-1][0]:
        return pts[-1][1]
    for (x0, s0), (x1, s1) in zip(pts, pts[1:]):
        if x0 <= x <= x1:
            return s0 + (s1 - s0) * (x - x0) / (x1 - x0)
    return pts[-1][1]


def _fmt_pct(x: Optional[float]) -> str:
    return "—" if x is None else f"{x*100:+.1f}%"


def _fmt_num(x: Optional[float], nd: int = 2) -> str:
    return "—" if x is None else f"{x:.{nd}f}"


# ==========================================================================
#   核心: 单标的基本面分析
# ==========================================================================
def analyze_one(name: str, code: str, market: str, is_etf: bool = False) -> FundamentalReport:
    """对单个股票/ETF 做基本面分析。yfinance 主干, 全程容错。"""
    if yf is None:
        return FundamentalReport(name, market, is_etf, False, "—", "", "", 0, "—", {}, [],
                                 "—", "", None, None, [], ["未安装 yfinance"],
                                 note="缺 yfinance, 无法分析")
    code_y = _norm(code)
    try:
        t = yf.Ticker(code_y)
        info = t.info or {}
    except Exception as e:
        return FundamentalReport(name, market, is_etf, False, "—", "", "", 0, "—", {}, [],
                                 "—", "", None, None, [], [f"信息抓取失败: {e}"])

    news = _fetch_news(t)

    # ETF / 指数型: 不做个股财报, 给一句话说明 (用户的深度因子是给个股的)
    qtype = (info.get("quoteType") or "").upper()
    if is_etf or qtype == "ETF" or "ETF" in name:
        return FundamentalReport(
            name, market, True, True, "指数/ETF", info.get("category", ""), "", 50.0, "—(指数型)",
            {}, [("类型", "指数/ETF", "基本面=成分股加权, 不做个股财报"),
                 ("跟踪", info.get("category", "—"), "关注指数整体估值分位与资金流")],
            "—", "一篮子分散, 无个股护城河概念", None, None, news,
            ["指数型: 深度财报因子不适用"], note="指数型标的, 看指数估值而非个股基本面")

    # --- 财报三表 (容错) --- #
    fin = _safe(lambda: t.financials)
    cf = _safe(lambda: t.cashflow)

    # 基础字段 (info 优先, 缺则财报兜底)
    rev = info.get("totalRevenue") or _get_row(fin, ["Total Revenue", "Operating Revenue"])
    rev_growth = info.get("revenueGrowth")
    earn_growth = info.get("earningsGrowth")
    gm = info.get("grossMargins")
    if gm is None:
        gp = _get_row(fin, ["Gross Profit"])
        gm = (gp / rev) if (gp and rev) else None
    opm = info.get("operatingMargins")
    npm = info.get("profitMargins")
    roe = info.get("returnOnEquity")
    d2e = info.get("debtToEquity")           # yfinance 多为百分数 (如 33.4 = 33%)
    ocf = info.get("operatingCashflow") or _get_row(cf, ["Operating Cash Flow", "Total Cash From Operating Activities"])
    fcf = info.get("freeCashflow")
    capex = _get_row(cf, ["Capital Expenditure", "Capital Expenditures"])
    if fcf is None and ocf is not None and capex is not None:
        fcf = ocf + capex                    # capex 在 yfinance 里通常为负值
    fcf_margin = (fcf / rev) if (fcf and rev) else None

    pe = info.get("trailingPE")
    pb = info.get("priceToBook")
    ps = info.get("priceToSalesTrailing12Months")
    divy = info.get("dividendYield")
    sector_raw = info.get("sector", "")
    industry_raw = info.get("industry", "")
    prof = _classify_sector(sector_raw, industry_raw)

    # 高阶/排雷因子 (财报里有则算, 没有则 N/A)
    sbc = _get_row(cf, ["Stock Based Compensation"])
    sbc_ratio = (sbc / rev) if (sbc and rev) else None
    rnd = _get_row(fin, ["Research And Development"])
    rnd_ratio = (rnd / rev) if (rnd and rev) else None
    # Magic Number ≈ 毛利增量 / 销售费用 (替代 LTV/CAC)
    gp_now = _get_row(fin, ["Gross Profit"])
    gp_prev = _get_row_prev(fin, ["Gross Profit"])
    sm = _get_row(fin, ["Selling And Marketing Expense", "Selling General And Administration"])
    magic = ((gp_now - gp_prev) / sm) if (gp_now and gp_prev and sm) else None

    # Rule of 40 = 营收增速% + FCF利润率%
    rule40 = None
    rule40_pass = None
    if rev_growth is not None and fcf_margin is not None:
        rule40 = (rev_growth + fcf_margin) * 100
        rule40_pass = rule40 >= 40.0

    # PEG (估值/成长性)
    peg = None
    if pe is not None and earn_growth and earn_growth > 0:
        peg = pe / (earn_growth * 100)

    flags: List[str] = []

    # ---------------- 分维度打分 (0~100) ---------------- #
    sub: Dict[str, float] = {}
    sub["增长"] = _piecewise(rev_growth, [(-0.10, 0), (0, 30), (0.15, 60), (0.30, 85), (0.50, 100)])
    sub["毛利"] = _piecewise(gm, [(prof.gross_margin_target - 0.25, 25),
                                  (prof.gross_margin_target, 65),
                                  (prof.gross_margin_target + 0.20, 95)])
    sub["FCF现金"] = _piecewise(fcf_margin, [(-0.10, 20), (0, 42), (0.10, 65), (0.20, 85), (0.30, 100)])
    sub["Rule40"] = _piecewise(rule40, [(10, 25), (40, 75), (60, 95)]) if rule40 is not None else None
    sub["ROE"] = _piecewise(roe, [(0, 20), (0.05, 35), (0.15, 70), (0.25, 95)])
    # 估值: 成长股用 PEG, 价值股用 PB
    if prof.is_growth and peg is not None:
        sub["估值"] = _piecewise(peg, [(0.5, 95), (1.0, 75), (2.0, 45), (3.0, 25)])
    else:
        sub["估值"] = _piecewise(pb, [(1.0, 90), (3.0, 65), (6.0, 40), (10.0, 20)])
    # 健康度: 负债率 (越低越好)
    sub["健康度"] = _piecewise(d2e, [(20, 90), (50, 70), (100, 45), (180, 20)]) if d2e is not None else None

    # SBC 排雷: >20% 重罚
    sbc_penalty = 0.0
    if sbc_ratio is not None and sbc_ratio > 0.20:
        sbc_penalty = min(15.0, (sbc_ratio - 0.20) * 100)
        flags.append(f"⚠️ SBC占营收 {sbc_ratio*100:.0f}% (>20%): 靠股权稀释撑 Non-GAAP 盈利, 扣 {sbc_penalty:.0f} 分")

    # 缺失项标注 + 加权 (缺的项剔除并重新归一)
    weights = {"增长": prof.w_growth, "毛利": prof.w_gm, "FCF现金": prof.w_fcf,
               "Rule40": prof.w_rule40, "ROE": prof.w_roe, "估值": prof.w_val, "健康度": prof.w_health}
    avail = {k: v for k, v in sub.items() if v is not None}
    miss = [k for k in sub if sub[k] is None]
    if miss:
        flags.append("数据缺失: " + "、".join(miss) + " (已按可得因子归一)")
    wsum = sum(weights[k] for k in avail) or 1.0
    score = sum(avail[k] * weights[k] for k in avail) / wsum
    score = max(0.0, score - sbc_penalty)
    grade = "A·优" if score >= 80 else "B·良" if score >= 65 else "C·中" if score >= 50 else "D·弱"

    # ---------------- 护城河: 行业定性档 + 量化加强 ---------------- #
    moat_tier, moat_reason = prof.moat_tier, prof.moat_reason
    if gm is not None and gm >= prof.gross_margin_target + 0.10 and (roe or 0) >= 0.18:
        moat_tier += "↑"
        moat_reason += f" (高毛利{gm*100:.0f}%+高ROE{(roe or 0)*100:.0f}%, 定价权强, 壁垒加强)"
    elif gm is not None and gm < prof.gross_margin_target - 0.10:
        moat_tier += "↓"
        moat_reason += f" (毛利{gm*100:.0f}%低于行业基准, 疑似定价权弱/竞争激烈)"

    # ---------------- 关键指标专栏 (用户要的"几个关键数字") ---------------- #
    def judge_gm():
        if gm is None: return "—"
        return "优" if gm >= prof.gross_margin_target + 0.1 else "达标" if gm >= prof.gross_margin_target else "偏低⚠️"
    metrics: List[Tuple[str, str, str]] = [
        ("营收增速(YoY)", _fmt_pct(rev_growth), "成长持续性" + ("强" if (rev_growth or 0) >= 0.2 else "一般")),
        ("毛利率", _fmt_pct(gm), f"{prof.key}基准{prof.gross_margin_target*100:.0f}% · {judge_gm()}"),
        ("FCF利润率", _fmt_pct(fcf_margin), "现金转换" + ("健康" if (fcf_margin or 0) >= 0.1 else "偏弱/烧钱" if fcf_margin is not None else "—")),
        ("Rule of 40", _fmt_num(rule40, 1) if rule40 is not None else "—",
         ("✅达标(增速+FCF≥40)" if rule40_pass else "❌未达标(易杀估值)") if rule40 is not None else "需增速+FCF"),
        ("ROE", _fmt_pct(roe), "股东回报" + ("优" if (roe or 0) >= 0.2 else "中" if (roe or 0) >= 0.1 else "弱")),
        ("净利率", _fmt_pct(npm), "—"),
        ("负债率(D/E)", _fmt_num(d2e, 1) + "%" if d2e is not None else "—",
         "稳健" if (d2e or 0) < 50 else "偏高⚠️"),
        ("估值 PE/PB/PS", f"{_fmt_num(pe,1)}/{_fmt_num(pb,1)}/{_fmt_num(ps,1)}",
         f"PEG={_fmt_num(peg,2)}" + (" 便宜" if (peg and peg < 1) else " 偏贵" if (peg and peg > 2) else "")),
        ("研发强度(R&D/营收)", _fmt_pct(rnd_ratio), "科技壁垒投入" if rnd_ratio is not None else "披露项"),
        ("SBC占营收", _fmt_pct(sbc_ratio), "排雷:>20%警惕稀释" if sbc_ratio is not None else "披露项/可手填"),
        ("Magic Number", _fmt_num(magic, 2), "销售转化效率(替代LTV/CAC)" if magic is not None else "需销售费用明细"),
        ("股息率", f"{divy:.2f}%" if divy is not None else "—", "—"),  # yfinance 已是百分数
    ]
    # 披露项 (免费源不可得, 手填则显示)
    ov = MANUAL_OVERRIDES.get(name, {})
    for label, key in [("Billings(订单出货)", "Billings"), ("RPO(剩余履约)", "RPO"),
                       ("ARR/MRR(经常性收入)", "ARR"), ("NRR/NDR(净留存)", "NRR"),
                       ("Guidance预期差", "Guidance")]:
        metrics.append((label, ov.get(key, "—"), "披露项·免费源不可得·可手填 MANUAL_OVERRIDES"))
    if ov.get("叙事"):
        metrics.append(("叙事/消息面(手填)", ov["叙事"], "用户定性补充"))

    return FundamentalReport(
        name, market, False, True, prof.key, sector_raw, industry_raw,
        round(score, 1), grade, {k: round(v, 1) for k, v in avail.items()},
        metrics, moat_tier, moat_reason, rule40, rule40_pass, news, flags,
    )


def _norm(code: str) -> str:
    """港股 4 位规范化 (与 ai.final 一致)。"""
    if isinstance(code, str) and code.endswith(".HK"):
        return code.replace(".HK", "").lstrip("0").zfill(4) + ".HK"
    return code


def _safe(fn):
    try:
        return fn()
    except Exception:
        return None


def analyze_fundamentals(specs: List[Tuple[str, str, str, bool]],
                         verbose: bool = True) -> Dict[str, FundamentalReport]:
    """
    批量分析 [(展示名, 代码, 市场, 是否ETF)] -> {展示名: FundamentalReport}。
    供 ai.final 融合调用 (币圈不传进来即可)。
    """
    out: Dict[str, FundamentalReport] = {}
    for name, code, market, is_etf in specs:
        if verbose:
            print(f"🔬 基本面分析 {name} ({code}) ...")
        rep = analyze_one(name, code, market, is_etf)
        out[name] = rep
        if verbose and rep.ok and not rep.is_etf:
            print(f"   {rep.grade} 分{rep.score} | {rep.sector_cn} | 护城河{rep.moat_tier} | "
                  f"Rule40={_fmt_num(rep.rule40,1) if rep.rule40 is not None else '—'}")
        time.sleep(0.4)
    return out


# ==========================================================================
#   Excel: 基本面专栏页 + 三支柱综合研判页
# ==========================================================================
def build_fundamental_sheet(wb, reports: Dict[str, FundamentalReport]) -> None:
    """在 Workbook 追加「📊 中长线基本面」页 —— 逐资产专栏, 关键指标全列, 不省略。"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    TITLE_FILL, SEC_FILL, HDR_FILL = "1F4E78", "A9D08E", "EDEDED"
    GRADE_FILL = {"A": "C6EFCE", "B": "E2EFDA", "C": "FFEB9C", "D": "FFC7CE"}

    ws = wb.create_sheet("📊 中长线基本面")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([24, 22, 36], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def merged3(r, text, fill=None, font=None, align=None, height=None, border=False):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(row=r, column=1, value=text)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        c.font = font or Font(size=11)
        c.alignment = align or LEFT
        if border:
            for col in range(1, 4):
                ws.cell(row=r, column=col).border = BORDER
        if height:
            ws.row_dimensions[r].height = height

    r = 1
    merged3(r, "📊 中长线基本面分析 · 四维因子库 (增长 / 现金流 / 效率粘性 / 前瞻) + 护城河 + 消息面",
            fill=TITLE_FILL, font=Font(bold=True, size=14, color="FFFFFF"), align=CENTER, height=30); r += 1
    merged3(r, "科技成长股按 增长持续性/现金流转换/业绩确定性 定价; 周期股按 估值/ROE/股息。"
               "披露项(Billings/RPO/ARR/NRR/Guidance)免费源不可得, 标注'可手填'。",
            font=Font(italic=True, size=9, color="555555"), height=24); r += 1
    r += 1

    for rep in reports.values():
        g = rep.grade[0]
        fill = GRADE_FILL.get(g, "F2F2F2")
        merged3(r, f"🏢 【{rep.name}】 {rep.market}   |   {rep.sector_cn}"
                   + (f" / {rep.industry_raw}" if rep.industry_raw else "")
                   + f"   |   基本面评级 {rep.grade} (分 {rep.score})",
                fill=fill, font=Font(bold=True, size=12, color="222222"), align=LEFT, height=26, border=True); r += 1

        if not rep.ok:
            merged3(r, f"   ⚠️ {rep.note or '财报数据抓取失败'}",
                    font=Font(italic=True, color="C00000"), height=20, border=True); r += 2
            continue
        if rep.is_etf:
            merged3(r, "   指数/ETF: 基本面=成分股加权, 不做个股财报。关注指数整体估值分位与资金流。",
                    font=Font(size=10), height=20, border=True); r += 2
            continue

        # 子项得分
        merged3(r, "  ① 分维度得分 (透明可审计)", fill=SEC_FILL,
                font=Font(bold=True, size=11, color="1F3D14"), height=20, border=True); r += 1
        sub_txt = "    " + "   ".join(f"{k}:{v:.0f}" for k, v in rep.subscores.items())
        merged3(r, sub_txt, font=Font(size=10), height=18); r += 1

        # 关键指标专栏 (三列: 指标 | 数值 | 评判)
        merged3(r, "  ② 关键指标专栏 (华尔街给科技股定价的核心因子)", fill=SEC_FILL,
                font=Font(bold=True, size=11, color="1F3D14"), height=20, border=True); r += 1
        for col, h in enumerate(["指标", "数值", "评判 / 说明"], 1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = Font(bold=True, size=10); c.fill = PatternFill("solid", fgColor=HDR_FILL)
            c.alignment = CENTER; c.border = BORDER
        r += 1
        for label, val, judge in rep.metrics:
            for col, v in enumerate([label, val, judge], 1):
                c = ws.cell(row=r, column=col, value=v)
                c.font = Font(size=10); c.border = BORDER
                c.alignment = LEFT if col != 2 else CENTER
            r += 1

        # 护城河 + 消息面 + 排雷
        merged3(r, "  ③ 护城河 / 消息面 / 排雷", fill=SEC_FILL,
                font=Font(bold=True, size=11, color="1F3D14"), height=20, border=True); r += 1
        merged3(r, f"    护城河: {rep.moat_tier} —— {rep.moat_reason}", font=Font(size=10), height=18); r += 1
        if rep.news:
            merged3(r, "    消息面(近一月):", font=Font(size=10, bold=True), height=16); r += 1
            for h in rep.news:
                merged3(r, f"      · {h}", font=Font(size=9, color="444444"), height=15); r += 1
        else:
            merged3(r, "    消息面: 暂无英文新闻流 (A股可参考持仓页东财股吧, 或手填叙事)",
                    font=Font(size=9, italic=True, color="777777"), height=15); r += 1
        for fl in rep.flags:
            merged3(r, f"    {fl}", font=Font(size=9, color="C00000"), height=15); r += 1

        ws.row_dimensions[r].height = 8
        r += 1


# --------------------------------------------------------------------------- #
# 三支柱综合研判 (技术 × 统计 × 基本面)
# --------------------------------------------------------------------------- #
def _tech_score(ana: Optional[dict]) -> Tuple[float, str]:
    """ai.final 技术研判 -> (0~100, 文字)。"""
    if not ana:
        return 50.0, "无技术数据"
    v = ana.get("verdict", "")
    if "偏多" in v:
        return 72.0, v
    if "偏空" in v:
        return 28.0, v
    return 50.0, v or "中性"


def _stat_score(dec) -> Tuple[float, str]:
    """midterm 统计加仓决策 -> (0~100, 文字)。"""
    if dec is None:
        return 50.0, "无统计数据"
    base = {"强烈加仓": 82, "可加仓": 66, "小幅试探": 55, "观望等待": 42, "暂不加仓": 30}
    s = base.get(dec.action, 50)
    p = dec.primary
    txt = dec.action + (f" (180天胜率{p.win_rate*100:.0f}%, 超额{p.edge_lo*100:+.0f}pp)" if p else "")
    return float(s), txt


def build_summary_sheet(wb, technical: Dict[str, dict], stat_decisions: list,
                        fundamentals: Dict[str, FundamentalReport]) -> None:
    """
    在 Workbook 追加「🎯 综合研判」页 —— 技术 / 统计 / 基本面【三支柱并列】+ 加权综合分 + 一句话结论。
    诚实: 三支柱常打架, 表里并列展示原貌, 综合分只作排序参考、不替代独立判断。
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    stat_by = {d.name: d for d in stat_decisions}
    names = list(dict.fromkeys(list(technical.keys()) + [d.name for d in stat_decisions]
                               + list(fundamentals.keys())))

    ws = wb.create_sheet("🎯 综合研判")
    ws.sheet_view.showGridLines = False
    headers = ["资产", "技术分析", "统计量化", "基本面", "综合分", "一句话结论"]
    widths = [20, 26, 30, 22, 9, 46]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "🎯 中长线综合研判 · 技术 × 统计量化 × 基本面 (三支柱并列, 综合分仅作排序参考)"
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1F4E78"); t.alignment = CENTER
    ws.row_dimensions[1].height = 28
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, size=10); c.fill = PatternFill("solid", fgColor="EDEDED")
        c.alignment = CENTER; c.border = BORDER

    rows = []
    for nm in names:
        ana = technical.get(nm)
        dec = stat_by.get(nm)
        fund = fundamentals.get(nm)
        is_crypto = ("BTC" in nm or "ETH" in nm or "比特币" in nm or "以太坊" in nm)

        ts, t_txt = _tech_score(ana)
        ss, s_txt = _stat_score(dec)
        if is_crypto or fund is None or not fund.ok or fund.is_etf:
            f_txt = "—(币圈不适用)" if is_crypto else ("指数型" if (fund and fund.is_etf) else "无数据")
            fs = None
            # 币圈/指数: 只用 技术×统计 各半
            composite = 0.5 * ts + 0.5 * ss
        else:
            f_txt = f"{fund.grade} (分{fund.score}, 护城河{fund.moat_tier})"
            fs = fund.score
            composite = 0.30 * ts + 0.30 * ss + 0.40 * fs

        # 一句话结论 (描述三者, 给净倾向)
        lean = ("积极加仓" if composite >= 70 else "可逢低分批" if composite >= 56
                else "中性持有" if composite >= 45 else "谨慎/观望")
        concl = f"{lean}: 技术{_lab(ts)}·统计{_lab(ss)}·基本面{(_lab(fs) if fs is not None else 'NA')}"
        if fs is not None and ts < 45 and fs >= 65:
            concl += " | 贵在便宜+基本面托底, 技术超卖区左侧机会"
        elif fs is not None and ts >= 60 and fs < 50:
            concl += " | 技术强但基本面弱, 防追高/逢强减"

        rows.append((nm, t_txt, s_txt, f_txt, round(composite, 0), concl, composite))

    rows.sort(key=lambda x: -x[6])
    rr = 3
    for nm, t_txt, s_txt, f_txt, comp, concl, _c in rows:
        for col, v in enumerate([nm, t_txt, s_txt, f_txt, comp, concl], 1):
            c = ws.cell(row=rr, column=col, value=v)
            c.font = Font(size=10); c.border = BORDER
            c.alignment = LEFT if col in (1, 2, 3, 4, 6) else CENTER
        fillc = "C6EFCE" if comp >= 70 else "E2EFDA" if comp >= 56 else "FFEB9C" if comp >= 45 else "FFC7CE"
        ws.cell(row=rr, column=5).fill = PatternFill("solid", fgColor=fillc)
        rr += 1
    rr += 1
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
    note = ws.cell(row=rr, column=1,
                   value="说明: 综合分=技术0.3+统计0.3+基本面0.4 (币圈/指数型按技术0.5+统计0.5)。"
                         "三支柱常打架, 请看并列原貌而非只看综合分。决策仅供参考, 不构成投资建议。")
    note.font = Font(italic=True, size=9, color="777777"); note.alignment = LEFT


def _lab(s: Optional[float]) -> str:
    if s is None:
        return "NA"
    return "🟢强" if s >= 65 else "🔴弱" if s < 40 else "🟡中"


# ==========================================================================
#   独立运行: 对 ai.final 持仓的股票部分跑基本面 (调试用)
# ==========================================================================
if __name__ == "__main__":
    DEMO = [
        ("瑞芯微", "603893.SS", "A股", False),
        ("腾讯控股", "0700.HK", "港股", False),
        ("中芯国际", "0981.HK", "港股", False),
    ]
    reps = analyze_fundamentals(DEMO)
    for nm, rep in reps.items():
        print("\n" + "=" * 60)
        print(f"【{nm}】 {rep.sector_cn} | {rep.grade} 分{rep.score} | 护城河{rep.moat_tier}")
        print(f"  Rule of 40: {rep.rule40}  达标={rep.rule40_pass}")
        print(f"  子项: {rep.subscores}")
        print("  关键指标:")
        for label, val, judge in rep.metrics:
            print(f"    {label:18}: {val:14} | {judge}")
        if rep.flags:
            print("  提示:", "; ".join(rep.flags))
