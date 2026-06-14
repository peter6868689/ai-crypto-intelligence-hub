# -*- coding: utf-8 -*-
"""
midterm_position_engine.py
================================
模块定位: 中长线复利机器 [加仓决策大脑] —— 用机构量化的统计纪律, 回答一个朴素问题:
          "我长期看好的资产, 此刻是不是一个值得加仓的好价格? 加仓胜率几何? 该加多少?"

它和这台机器里另外两套东西的血缘关系
---------------------------------------------------------------------------
  借鉴 [短线统计套利系统] (statarb_alpha_generator / paper_trading_engine / ...):
      不是搬它的协整数学 —— 配对均值回归那套数学只对"两条腿做价差"成立,
      搬到"单边做多加仓"上是错的, 不诚实。借鉴的是它的【方法论纪律】:
        1. 严格因果、零前视: 每个决策点只用"截至当下"的数据。
        2. 一切结论带样本量与置信区间 (Wilson), 小样本不许吹成铁律。
        3. Regime 状态机: 区分"均值回归的便宜" 与 "单边崩塌的下跌中继"。
        4. 用回测诚实证伪自己: 智能加仓 vs 无脑定投, 跑长历史比净值, 不行就承认。

  借鉴 [ai final.py] (全资产决策看板):
      它的四流派指标 (趋势 / 价值 / 酒田蜡烛 / 马丁回撤分档) + 情绪系统 (VIX /
      恐慌贪婪) + 日K量价图, 在这里【降级为"特征"】—— 不再各自拍脑袋喊多空,
      而是统一喂进统计引擎, 由历史频率说话。马丁的 -8/-15/-25% 分档被保留为
      "底仓档位", 再用统计胜率的置信度去缩放每一档实际加多少。

核心方法 (把"机构量化的统计严格性"真正落到加仓上的桥)
---------------------------------------------------------------------------
  不预测涨跌。对每个资产, 在每根日K上算一个【机会分 OpportunityScore】(越高=越便宜/
  越超卖, 由回撤深度 + RSI + 布林位置 + 价格相对长均线的 Z 值合成, 全部因果)。
  然后问历史: 过去每当机会分落在"当前这一档"时, 未来 N 天 (90/180/365, 对应你
  半年到一年的视角) 收益为正的比例是多少? —— 这就是【加仓胜率】, 配 Wilson 置信
  区间, 再减去"无条件基准胜率"得到真实 edge (它同时回答了"是否比无脑定投更值")。

  加仓仓位 = 马丁回撤档位的底仓比例 × 统计置信度乘数 (edge 显著且样本足 -> 加满该档;
  edge 不显著或样本稀 -> 缩水; Regime=单边破坏期 -> 进一步压低并示警)。

依赖:
    pip install yfinance ccxt pandas numpy
    # 可选 (缺失自动降级, 不崩): matplotlib mplfinance openpyxl requests
用法:
    python midterm_position_engine.py                 # 扫描全部资产, 出加仓决策报告
    python midterm_position_engine.py --asset 比特币    # 只看单个资产 (支持名字模糊匹配)
    python midterm_position_engine.py --backtest 比特币 # 智能加仓 vs 无脑定投, 长历史证伪
    python midterm_position_engine.py --excel          # 额外落一个桌面 Excel 看板
"""

from __future__ import annotations

import math
import os
import sys
import time
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

# ---------- 第三方库容错导入 (缺啥都不让整体崩) ----------
try:
    import yfinance as yf
except Exception:
    yf = None
try:
    import ccxt
except Exception:
    ccxt = None
try:
    import requests
except Exception:
    requests = None

pd.set_option("display.unicode.east_asian_width", True)


# ==========================================================================
#   全局配置: 资产宇宙 (与 ai final.py 对齐) + 决策参数
# ==========================================================================
# 股票 / ETF (代码, 市场标签) —— 走 yfinance 拉多年日K
STOCK_ASSETS: Dict[str, Tuple[str, str]] = {
    "瑞芯微":        ("603893.SS", "A股"),
    "艾为电子":      ("688798.SS", "A股"),
    "中芯国际":      ("0981.HK",   "港股"),
    "科创芯片ETF":   ("588780.SS", "A股"),
    "纳指ETF广发":   ("159941.SZ", "A股"),
    "标普500ETF":    ("513500.SS", "A股"),
    "恒生科技ETF":   ("513130.SS", "A股"),
    "黄金ETF华安":   ("518880.SS", "A股"),
    "腾讯控股":      ("0700.HK",   "港股"),
}
# 币圈资产 (欧易 OKX 现货符号) —— 走 ccxt 拉多年日K
CRYPTO_ASSETS: Dict[str, str] = {
    "比特币 BTC": "BTC/USDT",
    "以太坊 ETH": "ETH/USDT",
}

# --- 决策口径参数 (一处定义, 全局共用, 杜绝口径漂移) --- #
HORIZONS: Tuple[int, ...] = (90, 180, 365)   # 前向收益评估窗 (天): 半年到一年视角
PRIMARY_HORIZON: int = 180                   # 主口径: 用它定档与给结论
WARMUP_BARS: int = 252                        # 特征预热: 不足一年日K不给信号 (统计不可靠)
N_BUCKETS: int = 5                            # 机会分分桶数 (五分位)
MIN_BUCKET_SAMPLES: int = 20                  # 单桶最少样本, 不足则置信度强制打折
KNN_NEIGHBORS: int = 40                       # 最近历史相似日的个数 (二次交叉验证)
TXN_COST: float = 0.0015                      # 加仓单边综合成本 (费+滑点), 仅用于回测如实扣减

# 马丁回撤底仓档位 (沿用 ai.final 的 -8/-15/-25 思路, 多加一个浅档试探)
#   (从252日高点的回撤阈值, 档名, 该档底仓占"预留加仓弹药"的比例)
MARTINGALE_TIERS: List[Tuple[float, str, float]] = [
    (-0.25, "三档·深度回撤", 0.30),
    (-0.15, "二档·中度回调", 0.20),
    (-0.08, "首档·浅回调",   0.12),
    (-0.03, "试探·微回踩",   0.06),
]


# ==========================================================================
#   模块 1: 数据底座 (多年日K, 全程 try-except)
# ==========================================================================
def _normalize_hk(code: str) -> str:
    """港股代码统一成雅虎认的 4 位, 如 00981.HK -> 0981.HK。"""
    if isinstance(code, str) and code.endswith(".HK"):
        num = code.replace(".HK", "").lstrip("0")
        return num.zfill(4) + ".HK"
    return code


def fetch_stock_daily(code: str, years: int = 6) -> Optional[pd.DataFrame]:
    """yfinance 拉多年日K (越长统计样本越足), 返回标准 OHLCV, 失败 None。"""
    if yf is None:
        return None
    code = _normalize_hk(code)
    try:
        df = yf.download(code, period=f"{years}y", interval="1d",
                         auto_adjust=True, progress=False)  # 后复权, 干掉分红/拆分断崖
        if df is None or df.empty:
            return None
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        df = df[["Open", "High", "Low", "Close", "Volume"]].dropna()
        df.columns = ["open", "high", "low", "close", "volume"]
        return df
    except Exception as e:
        print(f"   ⚠️ 抓取 {code} 异常: {e}")
        return None


def fetch_crypto_daily(symbol: str, max_bars: int = 1500) -> Optional[pd.DataFrame]:
    """ccxt 欧易公开接口分页拉日K (突破单次 ~300 根上限), 返回标准 OHLCV, 失败 None。"""
    if ccxt is None:
        return None
    try:
        ex = ccxt.okx({"timeout": 15000, "enableRateLimit": True})
        tf_ms = ex.parse_timeframe("1d") * 1000
        now = ex.milliseconds()
        since = now - max_bars * tf_ms
        collected: List[list] = []
        while len(collected) < max_bars and since < now:
            batch = ex.fetch_ohlcv(symbol, timeframe="1d", since=since, limit=300)
            if not batch:
                break
            collected.extend(batch)
            nxt = batch[-1][0] + tf_ms
            if nxt <= since or len(batch) < 300:
                break
            since = nxt
        if not collected:
            return None
        dedup = {row[0]: row for row in collected}
        rows = [dedup[k] for k in sorted(dedup)][-max_bars:]
        df = pd.DataFrame(rows, columns=["t", "open", "high", "low", "close", "volume"])
        df["t"] = pd.to_datetime(df["t"], unit="ms")
        return df.set_index("t")[["open", "high", "low", "close", "volume"]]
    except Exception as e:
        print(f"   ⚠️ 欧易抓取 {symbol} 异常: {e}")
        return None


# ==========================================================================
#   模块 2: 情绪系统 (借自 ai final.py, 作为加仓的额外条件与示警)
# ==========================================================================
def fetch_crypto_fng() -> Tuple[Optional[int], str]:
    """币圈全网恐慌与贪婪指数, 返回 (数值, 中文描述)。极度恐慌往往是长线好价格。"""
    if requests is None:
        return None, "获取失败(缺 requests)"
    try:
        res = requests.get("https://api.alternative.me/fng/", timeout=6).json()
        v = int(res["data"][0]["value"])
        cls = res["data"][0]["value_classification"]
        return v, f"{v} ({cls})"
    except Exception:
        return None, "获取失败"


def fetch_vix() -> Tuple[Optional[float], str]:
    """VIX 恐慌指数 (股票资产的情绪锚)。>30 极度恐慌, 长线左侧可逢恐慌分批。"""
    if yf is None:
        return None, "数据缺省"
    try:
        df = yf.download("^VIX", period="5d", interval="1d",
                         auto_adjust=False, progress=False)
        if df is None or df.empty:
            return None, "数据缺省"
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        vix = float(df["Close"].iloc[-1])
        if vix > 30:
            return vix, f"{vix:.1f} 🚨极度恐慌(长线分批良机)"
        if vix > 20:
            return vix, f"{vix:.1f} ⚠️避险抬头"
        return vix, f"{vix:.1f} ⚪平稳"
    except Exception:
        return None, "数据缺省"


# ==========================================================================
#   模块 3: 特征工程 (ai.final 的指标 -> 因果特征序列)
# ==========================================================================
def _rsi(close: pd.Series, n: int = 14) -> pd.Series:
    delta = close.diff()
    gain = delta.clip(lower=0).rolling(n).mean()
    loss = (-delta.clip(upper=0)).rolling(n).mean()
    rs = gain / (loss + 1e-12)
    return 100 - 100 / (1 + rs)


def _zexpand(s: pd.Series, min_periods: int = WARMUP_BARS) -> pd.Series:
    """因果标准化: 用 expanding 均值/标准差, 每个点只用过去 -> 零前视。"""
    mu = s.expanding(min_periods=min_periods).mean()
    sd = s.expanding(min_periods=min_periods).std(ddof=1)
    return (s - mu) / sd.replace(0.0, np.nan)


def compute_features(df: pd.DataFrame) -> pd.DataFrame:
    """
    逐日算出加仓决策所需的全部【因果】特征。每个值只依赖"截至当日"的数据。

    机会分 opp_score: 越高 = 越便宜/越超卖 (越值得作为加仓候选)。由四个正交的
    "便宜信号"做因果 z 标准化后等权合成 —— 任何单一指标都可能骗人, 合成更稳健。
    """
    close, high, low, vol = df["close"], df["high"], df["low"], df["volume"]
    out = pd.DataFrame(index=df.index)
    out["close"] = close

    # --- ai.final 同源的核心指标 --- #
    out["rsi"] = _rsi(close, 14)
    ema12 = close.ewm(span=12, adjust=False).mean()
    ema26 = close.ewm(span=26, adjust=False).mean()
    dif = ema12 - ema26
    dea = dif.ewm(span=9, adjust=False).mean()
    out["macd_hist"] = (dif - dea) * 2

    ma20 = close.rolling(20).mean()
    sd20 = close.rolling(20).std(ddof=1)
    out["boll_pctb"] = (close - (ma20 - 2 * sd20)) / ((4 * sd20).replace(0.0, np.nan))  # 0=下轨,1=上轨

    out["ma50"] = close.rolling(50).mean()
    out["ma200"] = close.rolling(200).mean()
    out["trend_up"] = (out["ma50"] > out["ma200"]).astype(float)   # 长期多头排列?

    # 从过去 252 日滚动高点的回撤 (因果, 不含未来高点) —— 马丁分档的依据
    roll_high = high.rolling(252, min_periods=20).max()
    out["drawdown"] = close / roll_high - 1.0

    # 价格相对长均线的 Z 值 (借统计套利的"偏离度"思想, 但作用于单资产自身均衡)
    price_z = _zexpand(np.log(close) - np.log(close).rolling(120).mean())
    out["price_z"] = price_z

    # 量价: 近20日量能相对长期均量的比 (放量见底 / 缩量阴跌的刻画)
    out["vol_ratio"] = vol.rolling(20).mean() / (vol.rolling(120).mean() + 1e-12)

    # Regime (借短线系统): 高实现波动 或 高偏度 = 单边破坏期, 加仓需格外谨慎
    ret = np.log(close).diff()
    rvol = ret.rolling(30).std(ddof=1)
    rskew = ret.rolling(30).skew()
    vol_gate = rvol.expanding(min_periods=WARMUP_BARS).quantile(0.85)
    out["trending"] = ((rvol > vol_gate) | (rskew.abs() > 1.0)).astype(float)

    # --- 机会分: 四个"便宜信号"因果 z 标准化后等权合成 (越高越便宜) --- #
    comp_dd = _zexpand(-out["drawdown"])          # 回撤越深越高
    comp_rsi = _zexpand(50.0 - out["rsi"])        # RSI 越低越高
    comp_boll = _zexpand(0.5 - out["boll_pctb"])  # 越靠下轨越高
    comp_z = _zexpand(-out["price_z"])            # 价格越低于均衡越高
    out["opp_score"] = pd.concat([comp_dd, comp_rsi, comp_boll, comp_z], axis=1).mean(axis=1)

    return out


# ==========================================================================
#   模块 4: 统计核心 —— 条件化前向收益胜率 (带 Wilson 置信区间)
# ==========================================================================
def wilson_interval(wins: int, n: int, z: float = 1.96) -> Tuple[float, float, float]:
    """
    Wilson 二项比例置信区间 (比朴素 p±1.96·se 在小样本下稳健得多)。
    返回 (点估计 p, 下界, 上界)。n=0 时返回 (0.5, 0, 1) 表示"完全不知道"。
    """
    if n == 0:
        return 0.5, 0.0, 1.0
    p = wins / n
    denom = 1 + z * z / n
    center = (p + z * z / (2 * n)) / denom
    margin = (z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n))) / denom
    return p, max(0.0, center - margin), min(1.0, center + margin)


@dataclass(frozen=True)
class HorizonStat:
    """单个前向窗口的条件胜率统计快照。"""
    horizon: int
    n: int                  # 当前机会分档内的历史样本数
    win_rate: float         # 条件胜率点估计 P(前向收益>0 | 同档)
    win_lo: float           # Wilson 下界 (保守口径, 定仓用它)
    win_hi: float
    base_rate: float        # 无条件基准胜率 (该资产全样本)
    edge_lo: float          # win_lo - base_rate: 置信下界口径的真实超额
    mean_fwd: float         # 同档历史前向收益均值
    median_fwd: float


@dataclass(frozen=True)
class ConditionalStats:
    """某资产在"当前机会分档"下, 跨多个前向窗口的完整统计画像。"""
    bucket: int                         # 当前落入的机会分桶 (0=最贵, N-1=最便宜)
    bucket_label: str
    opp_score: float
    opp_pct: float                      # 当前机会分在历史中的分位 (0~1, 越高越便宜)
    per_horizon: Dict[int, HorizonStat]
    knn_win_rate: Dict[int, float]      # 最近 K 个相似历史日的前向胜率 (二次交叉验证)
    knn_n: int


def _bucketize(opp: pd.Series, valid: pd.Series, n_buckets: int) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    用【估计集】(有已实现前向收益的那部分历史) 的分位切桶边界, 把整条 opp_score 映射到桶号。
    返回 (桶号数组, 桶边界, 估计集掩码)。当前点用同一套边界映射, 避免口径漂移。
    """
    est = opp[valid].dropna()
    if len(est) < n_buckets * 5:
        edges = np.array([])
        return np.full(len(opp), -1), edges, valid.to_numpy()
    qs = np.linspace(0, 1, n_buckets + 1)
    edges = np.unique(np.quantile(est.to_numpy(), qs))
    # 桶号: digitize 到内部边界 (clip 到 [0, n_buckets-1])
    buckets = np.clip(np.digitize(opp.to_numpy(), edges[1:-1]), 0, len(edges) - 2)
    return buckets, edges, valid.to_numpy()


def conditional_stats(feat: pd.DataFrame) -> Optional[ConditionalStats]:
    """
    引擎心脏: 对一个资产, 计算"当前机会分档"下的条件化前向收益胜率画像。

    严格因果: 前向收益是标签 (天然向后看 horizon 天), 但只在【已能完整观测到 horizon
    天之后】的历史点上参与统计 (估计集); 当前决策点的前向收益未知, 绝不使用, 只用它的
    机会分去查"历史上同档的命运分布"。机会分本身全部因果 (见 compute_features)。
    """
    if feat is None or feat["opp_score"].notna().sum() < WARMUP_BARS:
        return None
    close = feat["close"]
    opp = feat["opp_score"]
    n = len(feat)

    # 当前点 (最后一根) 的机会分必须有效
    opp_now = opp.iloc[-1]
    if pd.isna(opp_now):
        return None

    # 用最大前向窗划定"可作估计样本"的范围: 这些点之后还有完整 horizon 天数据
    max_h = max(HORIZONS)
    base_valid = pd.Series(np.arange(n) < (n - max_h), index=feat.index) & opp.notna()
    buckets, edges, _ = _bucketize(opp, base_valid, N_BUCKETS)
    if edges.size == 0:
        return None
    cur_bucket = int(buckets[-1])
    n_real = len(edges) - 1
    label = ["最贵区", "偏贵区", "中性区", "偏便宜区", "最便宜区"]
    bucket_label = label[min(cur_bucket * len(label) // max(1, n_real), len(label) - 1)] \
        if n_real != len(label) else label[cur_bucket]
    opp_pct = float((opp[base_valid] < opp_now).mean()) if base_valid.any() else float("nan")

    per_h: Dict[int, HorizonStat] = {}
    for h in HORIZONS:
        fwd = close.shift(-h) / close - 1.0
        # 该窗口的合法估计点: 能观测到未来 h 天 且 机会分有效
        valid_h = (pd.Series(np.arange(n) < (n - h), index=feat.index)) & opp.notna() & fwd.notna()
        v = valid_h.to_numpy()
        in_bucket = (buckets == cur_bucket) & v
        fwd_v = fwd.to_numpy()
        # 条件样本
        cond = fwd_v[in_bucket]
        n_c = int(cond.size)
        wins_c = int((cond > 0).sum())
        p, lo, hi = wilson_interval(wins_c, n_c)
        # 无条件基准
        base = fwd_v[v]
        base_rate = float((base > 0).mean()) if base.size else 0.5
        per_h[h] = HorizonStat(
            horizon=h, n=n_c, win_rate=p, win_lo=lo, win_hi=hi,
            base_rate=base_rate, edge_lo=lo - base_rate,
            mean_fwd=float(cond.mean()) if n_c else float("nan"),
            median_fwd=float(np.median(cond)) if n_c else float("nan"),
        )

    # --- 二次交叉验证: 最近 K 个"特征最相似"历史日的前向胜率 (与分桶法相互印证) --- #
    knn_feats = ["drawdown", "rsi", "boll_pctb", "price_z"]
    fz = feat[knn_feats].apply(lambda s: (s - s.mean()) / (s.std(ddof=1) + 1e-12))
    cur_vec = fz.iloc[-1].to_numpy()
    knn_win: Dict[int, float] = {}
    knn_n = 0
    for h in HORIZONS:
        fwd = close.shift(-h) / close - 1.0
        mask = (pd.Series(np.arange(n) < (n - h), index=feat.index)) & fz.notna().all(axis=1) & fwd.notna()
        cand = fz[mask]
        if len(cand) < KNN_NEIGHBORS:
            knn_win[h] = float("nan")
            continue
        dist = np.sqrt(((cand.to_numpy() - cur_vec) ** 2).sum(axis=1))
        nn_idx = cand.index[np.argsort(dist)[:KNN_NEIGHBORS]]
        nn_fwd = fwd.loc[nn_idx]
        knn_win[h] = float((nn_fwd > 0).mean())
        knn_n = len(nn_idx)

    return ConditionalStats(
        bucket=cur_bucket, bucket_label=bucket_label, opp_score=float(opp_now),
        opp_pct=opp_pct, per_horizon=per_h, knn_win_rate=knn_win, knn_n=knn_n,
    )


# ==========================================================================
#   模块 5: 加仓决策引擎 (统计胜率 + 马丁档位 + 情绪/Regime 闸门)
# ==========================================================================
@dataclass(frozen=True)
class AddDecision:
    """单资产的加仓决策快照 —— 机器"此刻该不该加、加多少、凭什么"的可审计记录。"""
    name: str
    market: str
    price: float
    drawdown: float
    trending: bool
    opp_pct: float
    tier_name: str
    tier_base_pct: float            # 该回撤档的底仓比例 (占预留弹药)
    confidence: float               # 统计置信度乘数 (0.3~1.5)
    add_pct: float                  # 最终建议加仓比例 = 档位 × 置信度
    action: str                     # 强烈加仓 / 可加仓 / 小幅试探 / 观望等待 / 暂不加仓
    primary: Optional[HorizonStat]  # 主口径 (180天) 的统计快照
    stats: Optional[ConditionalStats]
    sentiment: str
    reasons: List[str] = field(default_factory=list)


def _tier_for(drawdown: float) -> Tuple[str, float]:
    """按从高点的回撤深度定马丁底仓档位 (越深档位越重)。无回撤 -> 0 档。"""
    for thr, name, pct in MARTINGALE_TIERS:
        if drawdown <= thr:
            return name, pct
    return "无回撤·高位", 0.0


def evaluate_add(name: str, market: str, df: pd.DataFrame,
                 sentiment_val: Optional[float], sentiment_str: str,
                 is_crypto: bool) -> AddDecision:
    """对单个资产产出完整加仓决策。统计说话, 情绪/Regime 作闸门与微调。"""
    feat = compute_features(df)
    stats = conditional_stats(feat)
    last = feat.iloc[-1]
    price = float(last["close"])
    dd = float(last["drawdown"]) if pd.notna(last["drawdown"]) else 0.0
    trending = bool(last["trending"]) if pd.notna(last["trending"]) else False

    tier_name, tier_base = _tier_for(dd)
    reasons: List[str] = []

    # 数据不足兜底
    if stats is None:
        reasons.append("历史样本不足 (需 >=1 年干净日K), 统计不可靠, 仅作观望。")
        return AddDecision(name, market, price, dd, trending,
                           float("nan"), tier_name, tier_base, 0.3, 0.0,
                           "观望等待", None, None, sentiment_str, reasons)

    primary = stats.per_horizon[PRIMARY_HORIZON]

    # --- 置信度乘数: 以"主口径胜率的 Wilson 下界 相对 基准"的超额定调 --- #
    # edge_lo 落在 [-0.15, +0.15] 映射到乘数 [0.5, 1.5]; 再受样本量/Regime/情绪修正。
    edge = primary.edge_lo
    conf = float(np.clip(1.0 + (edge / 0.15) * 0.5, 0.5, 1.5))

    if primary.n < MIN_BUCKET_SAMPLES:
        conf = min(conf, 0.7)
        reasons.append(f"当前档历史样本仅 {primary.n} 个 (<{MIN_BUCKET_SAMPLES}), 置信度打折。")

    if trending:
        conf = min(conf, 0.7)
        reasons.append("Regime=单边破坏期 (高波动/高偏度): 可能是下跌中继而非便宜, 压低仓位、分批为上。")

    # 情绪闸门: 极度恐慌往往是长线好价格 -> 轻微上调 (但不超过 1.5)
    fear_boost = False
    if is_crypto and sentiment_val is not None and sentiment_val <= 25:
        conf = min(conf * 1.15, 1.5); fear_boost = True
    if (not is_crypto) and sentiment_val is not None and sentiment_val >= 30:  # VIX
        conf = min(conf * 1.15, 1.5); fear_boost = True
    if fear_boost:
        reasons.append(f"市场情绪处于恐慌区 ({sentiment_str}) -> 长线左侧, 置信度上调。")

    add_pct = tier_base * conf

    # --- 行动分级 (诚实: 不给非黑即白的硬指令, 给带置信的建议) --- #
    p = primary.win_rate
    lo = primary.win_lo
    base = primary.base_rate
    if tier_base == 0.0:
        action = "暂不加仓"
        reasons.insert(0, f"距高点仅 {dd*100:+.1f}%, 仍在高位, 不在任何加仓档 -> 等回调。")
    elif lo >= base and lo >= 0.5 and primary.n >= MIN_BUCKET_SAMPLES and not trending:
        action = "强烈加仓"
        reasons.insert(0, f"{PRIMARY_HORIZON}天条件胜率 {p*100:.0f}% (下界 {lo*100:.0f}%) 显著跑赢基准 "
                          f"{base*100:.0f}%, 且非破坏期 -> 高置信加仓。")
    elif lo >= base and primary.n >= MIN_BUCKET_SAMPLES:
        action = "可加仓"
        reasons.insert(0, f"{PRIMARY_HORIZON}天条件胜率下界 {lo*100:.0f}% ≥ 基准 {base*100:.0f}%, "
                          f"统计上确有超额 -> 按档加仓。")
    elif p >= base:
        action = "小幅试探"
        reasons.insert(0, f"条件胜率点估计 {p*100:.0f}% 略高于基准 {base*100:.0f}% 但置信不足 "
                          f"(下界 {lo*100:.0f}%<基准) -> 仅小幅试探。")
    else:
        action = "观望等待"
        add_pct = min(add_pct, tier_base * 0.4)
        reasons.insert(0, f"条件胜率 {p*100:.0f}% 未跑赢基准 {base*100:.0f}%, "
                          f"当前便宜程度不构成统计优势 -> 观望或极小仓。")

    # kNN 印证
    knn = stats.knn_win_rate.get(PRIMARY_HORIZON)
    if knn is not None and not math.isnan(knn):
        agree = "印证" if (knn >= 0.5) == (p >= 0.5) else "分歧⚠️"
        reasons.append(f"最近 {stats.knn_n} 个相似历史日的 {PRIMARY_HORIZON}天胜率 {knn*100:.0f}% ({agree}分桶结论)。")

    reasons.append(f"机会分历史分位 {stats.opp_pct*100:.0f}% ({stats.bucket_label}), "
                   f"回撤 {dd*100:+.1f}% -> {tier_name}。")

    return AddDecision(name, market, price, dd, trending, stats.opp_pct,
                       tier_name, tier_base, round(conf, 3), round(add_pct, 4),
                       action, primary, stats, sentiment_str, reasons)


# ==========================================================================
#   模块 6: 回测验证 —— 智能加仓 vs 无脑定投 (诚实证伪)
# ==========================================================================
def backtest_dca_vs_smart(df: pd.DataFrame, name: str,
                          invest_every: int = 30, budget_per: float = 100.0) -> None:
    """
    长历史上对比两种加仓法的资金效率 (严格因果, 每次只用截至当日的信息):
        A) 无脑定投: 每 invest_every 天固定投 budget_per。
        B) 智能加仓: 同样的总预算与节奏, 但把每期资金按"当时机会分桶的条件胜率与回撤档"
                     动态分配 —— 便宜且统计占优时多投, 贵或无优势时少投/攒着。
    比的是: 同样的钱, 谁的期末持仓市值/成本更优 (即"好价格"是否真带来超额)。

    诚实声明: 这是"事后用全样本切桶"的近似验证 (估计集与决策集有重叠), 用于看方向性,
    不等于实盘无前视收益。它的价值在于: 若连这种宽松验证都跑不赢无脑定投, 智能加仓就是自欺。
    """
    feat = compute_features(df)
    feat = feat.dropna(subset=["opp_score", "drawdown"])
    if len(feat) < WARMUP_BARS + 60:
        print(f"   ⚠️ {name} 历史不足, 跳过回测。")
        return

    close = feat["close"]
    opp = feat["opp_score"]
    # 全样本五分位边界 (近似; 仅用于方向性验证)
    edges = np.unique(np.quantile(opp.to_numpy(), np.linspace(0, 1, N_BUCKETS + 1)))
    bucket = np.clip(np.digitize(opp.to_numpy(), edges[1:-1]), 0, len(edges) - 2)

    # 智能权重: 越便宜的桶权重越高 (1.0 -> N), 体现"好价格多投"
    weight_by_bucket = {b: 0.5 + b for b in range(len(edges) - 1)}

    idxs = list(range(0, len(feat), invest_every))
    total_budget = len(idxs) * budget_per

    # A) 无脑定投
    units_a = 0.0
    for i in idxs:
        px = float(close.iloc[i]) * (1 + TXN_COST)
        units_a += budget_per / px
    # B) 智能加仓: 同样总预算, 按机会分权重重新分配
    wsum = sum(weight_by_bucket[int(bucket[i])] for i in idxs)
    units_b = 0.0
    for i in idxs:
        alloc = total_budget * (weight_by_bucket[int(bucket[i])] / wsum)
        px = float(close.iloc[i]) * (1 + TXN_COST)
        units_b += alloc / px

    last_px = float(close.iloc[-1])
    val_a = units_a * last_px
    val_b = units_b * last_px
    cost_a = cost_b = total_budget
    print("\n" + "=" * 60)
    print(f"  回测验证 · 智能加仓 vs 无脑定投 | {name} | {len(feat)} 根日K")
    print("=" * 60)
    print(f"  总投入预算    : {total_budget:,.0f}  (每 {invest_every} 天投 {budget_per:.0f}, 共 {len(idxs)} 期)")
    print(f"  A 无脑定投市值 : {val_a:,.1f}   收益率 {(val_a/cost_a-1)*100:+.1f}%   均价 {cost_a/units_a:,.4f}")
    print(f"  B 智能加仓市值 : {val_b:,.1f}   收益率 {(val_b/cost_b-1)*100:+.1f}%   均价 {cost_b/units_b:,.4f}")
    edge = (val_b / val_a - 1) * 100
    verdict = "✅ 智能加仓跑赢 (好价格确有超额)" if val_b > val_a else "⚠️ 本样本未跑赢, 别迷信择时"
    print("-" * 60)
    print(f"  智能相对定投超额 : {edge:+.1f}%   结论: {verdict}")
    print("=" * 60)


# ==========================================================================
#   模块 7: 报告渲染 (控制台)
# ==========================================================================
def print_decision(d: AddDecision) -> None:
    icon = {"强烈加仓": "🟢🔥", "可加仓": "🟢", "小幅试探": "🟡",
            "观望等待": "⚪", "暂不加仓": "🔴", }.get(d.action, "•")
    print("\n" + "─" * 66)
    print(f" 【{d.name}】 {d.market}   现价 {d.price:,.4f}")
    print(f" {icon} 决策: {d.action}   建议加仓 {d.add_pct*100:.1f}% 弹药"
          f" (档位 {d.tier_base_pct*100:.0f}% × 置信 {d.confidence:.2f})")
    if d.primary is not None:
        s = d.primary
        print(f" 主口径 {s.horizon}天: 条件胜率 {s.win_rate*100:.0f}% "
              f"[{s.win_lo*100:.0f}~{s.win_hi*100:.0f}%]  基准 {s.base_rate*100:.0f}%  "
              f"超额下界 {s.edge_lo*100:+.0f}pp  n={s.n}  均涨 {s.mean_fwd*100:+.1f}%")
        # 其余窗口一并列出
        if d.stats is not None:
            parts = []
            for h in HORIZONS:
                hs = d.stats.per_horizon[h]
                parts.append(f"{h}天 {hs.win_rate*100:.0f}%(n{hs.n})")
            print(f" 多窗胜率: " + "  ".join(parts))
    print(f" 情绪: {d.sentiment}")
    for r in d.reasons:
        print(f"   ▪ {r}")
    print("─" * 66)


# ==========================================================================
#   模块 8: 主引擎
# ==========================================================================
def _match_asset(query: str) -> List[Tuple[str, str, str, bool]]:
    """名字模糊匹配, 返回 [(展示名, 代码/符号, 市场, 是否币圈)]。"""
    out = []
    for name, (code, mkt) in STOCK_ASSETS.items():
        if query in name or query in code:
            out.append((name, code, mkt, False))
    for name, sym in CRYPTO_ASSETS.items():
        if query in name or query in sym:
            out.append((name, sym, "币圈", True))
    return out


def _all_assets() -> List[Tuple[str, str, str, bool]]:
    out = [(n, c, m, False) for n, (c, m) in STOCK_ASSETS.items()]
    out += [(n, s, "币圈", True) for n, s in CRYPTO_ASSETS.items()]
    return out


ACTION_RANK: Dict[str, int] = {"强烈加仓": 0, "可加仓": 1, "小幅试探": 2,
                               "观望等待": 3, "暂不加仓": 4}


def collect_decisions(
    assets: List[Tuple[str, str, str, bool]],
    fng: Optional[Tuple[Optional[float], str]] = None,
    vix: Optional[Tuple[Optional[float], str]] = None,
    verbose: bool = True,
) -> List[AddDecision]:
    """
    对一组资产 [(展示名, 代码/符号, 市场, 是否币圈)] 抓多年日K + 评估, 返回已排序的决策列表。

    情绪 (恐慌贪婪 / VIX) 只拉一次全局共用。这是给【外部 (如 ai final.py)】融合调用的稳定入口:
    它不打印汇总、不出 Excel, 只负责"把决策算出来"。
    """
    if fng is None:
        fng = fetch_crypto_fng()
    if vix is None:
        vix = fetch_vix()
    fng_val, fng_str = fng
    vix_val, vix_str = vix

    decisions: List[AddDecision] = []
    for name, code, mkt, is_crypto in assets:
        if verbose:
            print(f"\n🔍 拉取 {name} ({code}) ...")
        df = fetch_crypto_daily(code) if is_crypto else fetch_stock_daily(code)
        if df is None or len(df) < WARMUP_BARS:
            if verbose:
                print(f"   ❌ {name} 数据不足或抓取失败 (需 >= {WARMUP_BARS} 根日K), 跳过。")
            continue
        sval, sstr = (fng_val, fng_str) if is_crypto else (vix_val, vix_str)
        d = evaluate_add(name, mkt, df, sval, sstr, is_crypto)
        if verbose:
            print_decision(d)
        decisions.append(d)
        time.sleep(0.3)

    decisions.sort(key=lambda x: (ACTION_RANK.get(x.action, 9), -x.add_pct))
    return decisions


# ----------------------------------------------------------------------------- #
# 组合层: 把"总加仓弹药"在各资产间分配 (借短线系统的分散纪律: 不把身家压一注)
# ----------------------------------------------------------------------------- #
def _cap_redistribute(shares: Dict[int, float], cap: float) -> Dict[int, float]:
    """把超过 cap 的份额削平, 溢出按比例再分给未封顶者 (满仓部署情形用)。"""
    s = dict(shares)
    for _ in range(20):
        over = {k: v for k, v in s.items() if v > cap + 1e-9}
        if not over:
            break
        excess = sum(v - cap for v in over.values())
        for k in over:
            s[k] = cap
        under = {k: v for k, v in s.items() if v < cap - 1e-9}
        usum = sum(under.values())
        if usum <= 1e-12 or excess <= 1e-12:
            break
        for k in under:
            s[k] += excess * (under[k] / usum)
    return s


def allocate_portfolio(
    decisions: List[AddDecision],
    reserve_capital: Optional[float] = None,
    max_single: float = 0.35,
) -> Tuple[List[dict], float]:
    """
    组合层分配 (诚实、可解释的口径):
      - 仅 {强烈加仓 / 可加仓 / 小幅试探} 的资产参与分配; 观望 / 暂不 -> 0。
      - 每个资产的 add_pct 直接解释为"建议投入占【总弹药】的比例"。
      - Σadd_pct ≤ 1 : 原样部署, 剩余 (1-Σ) 留作【干火药】(信心不足就别全压上, 留子弹)。
      - Σadd_pct > 1 : 归一到满仓, 按信心 (add_pct) 比例分配。
      - 任一资产封顶 max_single (默认 35%): 满仓情形溢出再分给未封顶者; 非满仓情形溢出转干火药。
    :return: (allocations, dry_powder_frac)。allocations 每项含 name/market/action/add_pct/share/amount。
    """
    qual = [d for d in decisions
            if d.add_pct > 0 and d.action in ("强烈加仓", "可加仓", "小幅试探")]
    raw = {id(d): d.add_pct for d in qual}
    total = sum(raw.values())

    if total <= 0:
        shares: Dict[int, float] = {}
    elif total > 1.0:
        shares = _cap_redistribute({k: v / total for k, v in raw.items()}, max_single)
    else:
        shares = {k: min(v, max_single) for k, v in raw.items()}

    allocations: List[dict] = []
    deployed = 0.0
    for d in decisions:
        s = float(shares.get(id(d), 0.0))
        deployed += s
        allocations.append({
            "name": d.name, "market": d.market, "action": d.action,
            "add_pct": d.add_pct, "share": s,
            "amount": (reserve_capital * s) if reserve_capital else None,
        })
    return allocations, max(0.0, 1.0 - deployed)


def scan(assets: List[Tuple[str, str, str, bool]], want_excel: bool = False,
         reserve_capital: Optional[float] = None) -> List[AddDecision]:
    print("🧭 中长线复利机器 · 加仓决策引擎 (半年~一年视角)")
    print("   统计纪律借自短线统计套利系统; 指标/情绪/出图借自 ai final.py")

    decisions = collect_decisions(assets, verbose=True)

    # 汇总: 按"可加仓优先 + 建议仓位降序"排序, 一眼看到当下最该加的
    print("\n" + "=" * 66)
    print("  📋 加仓优先级汇总 (越靠前越值得现在加)")
    print("=" * 66)
    for d in decisions:
        p = f"{d.primary.win_rate*100:.0f}%" if d.primary else "—"
        print(f"  {d.action:<6} | {d.name:<14} 现价{d.price:>12,.4f} "
              f"回撤{d.drawdown*100:>+6.1f}% 胜率{p:>4} -> 加 {d.add_pct*100:>4.1f}%")

    # 组合层: 总弹药分配
    allocations, dry = allocate_portfolio(decisions, reserve_capital)
    print("\n  💰 组合弹药分配 (按信心分配, 单资产封顶 35%, 信心不足留干火药)")
    any_alloc = False
    for a in allocations:
        if a["share"] <= 1e-6:
            continue
        any_alloc = True
        amt = f"  ≈ {a['amount']:,.0f}" if a["amount"] is not None else ""
        print(f"   {a['name']:<14} {a['action']:<6} 占总弹药 {a['share']*100:4.1f}%{amt}")
    if not any_alloc:
        print("   (当前无任何资产达到加仓条件 -> 全部留作干火药, 等更好的价格)")
    print(f"   🧱 干火药(留存观望) {dry*100:.1f}%"
          + (f"  ≈ {reserve_capital*dry:,.0f}" if reserve_capital else ""))
    print("=" * 66)
    print("  ⚠️ 决策仅供参考: 统计是概率不是承诺, 极端行情下历史规律可能失效。")

    if want_excel:
        try:
            _write_excel(decisions, reserve_capital)
        except Exception as e:
            print(f"   ⚠️ Excel 生成失败 (不影响控制台结论): {e}")
    return decisions


def build_midterm_sheets(wb, decisions: List[AddDecision],
                         reserve_capital: Optional[float] = None) -> None:
    """
    在【已有的 openpyxl Workbook】中追加中长线加仓决策页 —— 报告式纵向排版, 逐资产把
    所有字段【完整列出, 绝不省略】(这是 ai final.py 看板融合调用的入口)。

    共两页:
      Sheet「🧭 中长线加仓决策」: 每个资产一个完整区块 ①当前状态 ②加仓胜率全表(三窗×全字段)
                                  ③收益分布+kNN交叉验证 ④仓位建议拆解 ⑤全部决策依据。
      Sheet「💰 组合弹药分配」  : 组合层按信心把总弹药分到各资产 + 干火药留存。

    缺 openpyxl 时抛 ImportError, 由调用方兜底 (不让整体崩)。
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ACT_FILL = {"强烈加仓": "C6EFCE", "可加仓": "E2EFDA", "小幅试探": "FFEB9C",
                "观望等待": "F2F2F2", "暂不加仓": "FFC7CE"}
    SEC_FILL, HDR_FILL, TITLE_FILL = "A9D08E", "EDEDED", "1F4E78"

    # ====================== Sheet 1: 完整加仓决策报告 ====================== #
    ws = wb.create_sheet("🧭 中长线加仓决策")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([20, 14, 14, 14, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def merged(r, text, fill=None, font=None, align=None, height=None, border=False):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=text)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        c.font = font or Font(size=11)
        c.alignment = align or LEFT
        if border:
            for col in range(1, 7):
                ws.cell(row=r, column=col).border = BORDER
        if height:
            ws.row_dimensions[r].height = height

    def tablerow(r, vals, header=False):
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = BORDER
            c.alignment = CENTER
            if header:
                c.font = Font(bold=True, size=10)
                c.fill = PatternFill("solid", fgColor=HDR_FILL)
            else:
                c.font = Font(size=10)

    r = 1
    merged(r, "🧭 中长线加仓决策引擎 · 完整报告 (半年~一年视角 · 多年日K · 统计口径)",
           fill=TITLE_FILL, font=Font(bold=True, size=14, color="FFFFFF"),
           align=CENTER, height=30); r += 1
    merged(r, f"口径: 主窗 {PRIMARY_HORIZON} 天 | 胜率=历史上同档机会分之后该窗收益>0 的频率 (Wilson 95% 区间) | "
              f"超额下界=条件胜率下界−无条件基准 (>0 才算统计占优) | 仓位=马丁回撤档位 × 统计置信乘数",
           font=Font(italic=True, size=9, color="555555"), height=26); r += 1
    r += 1

    for d in decisions:
        fill = ACT_FILL.get(d.action, "FFFFFF")
        # ① 标题条
        merged(r, f"📊 【{d.name}】 {d.market}    现价 {d.price:,.4f}    →  决策【{d.action}】"
                  f"    建议加仓 {d.add_pct*100:.1f}% 弹药 (档位 {d.tier_base_pct*100:.0f}% × 置信 {d.confidence:.2f})",
               fill=fill, font=Font(bold=True, size=12, color="222222"),
               align=LEFT, height=28, border=True); r += 1

        if d.primary is None or d.stats is None:
            merged(r, "   ⚠️ 历史样本不足 (需 >= 1 年干净日K), 统计不可靠, 仅作观望。",
                   font=Font(italic=True, color="C00000"), height=20, border=True)
            r += 2
            continue

        st = d.stats
        # ① 当前状态
        merged(r, "①  当前状态", fill=SEC_FILL, font=Font(bold=True, size=11, color="1F3D14"),
               height=20, border=True); r += 1
        merged(r, f"      回撤(距 252 日高点): {d.drawdown*100:+.1f}%     "
                  f"机会分历史分位: {d.opp_pct*100:.0f}% ({st.bucket_label})     "
                  f"机会分值: {st.opp_score:+.2f}", font=Font(size=10), height=18); r += 1
        merged(r, f"      市场微观状态: {'⚠️ 单边破坏期 (高波动/高偏度, 慎防下跌中继)' if d.trending else '✅ 均值回归区'}"
                  f"     回撤档位: {d.tier_name}     情绪: {d.sentiment}",
               font=Font(size=10), height=18); r += 1

        # ② 加仓胜率全表 (三窗 × 核心字段)
        merged(r, "②  加仓胜率 · 条件化前向收益 (全部窗口, 含 95% 置信区间, 不省略)",
               fill=SEC_FILL, font=Font(bold=True, size=11, color="1F3D14"),
               height=20, border=True); r += 1
        tablerow(r, ["前向窗口", "样本 n", "条件胜率%", "95%下界%", "无条件基准%", "超额下界 pp"],
                 header=True); r += 1
        for h in HORIZONS:
            hs = st.per_horizon[h]
            tablerow(r, [f"{h} 天", hs.n, round(hs.win_rate*100, 1), round(hs.win_lo*100, 1),
                         round(hs.base_rate*100, 1), round(hs.edge_lo*100, 1)]); r += 1

        # ③ 收益分布 + kNN 交叉验证
        merged(r, "③  历史收益分布 + 相似历史日交叉验证 (kNN, 与分桶法相互印证)",
               fill=SEC_FILL, font=Font(bold=True, size=11, color="1F3D14"),
               height=20, border=True); r += 1
        tablerow(r, ["前向窗口", "历史均涨%", "历史中位%", "95%上界%", "kNN相似日胜率%", "与分桶"],
                 header=True); r += 1
        for h in HORIZONS:
            hs = st.per_horizon[h]
            knn = st.knn_win_rate.get(h)
            if knn is None or math.isnan(knn):
                knn_s, agree = "—", "—"
            else:
                knn_s = round(knn*100, 1)
                agree = "印证" if (knn >= 0.5) == (hs.win_rate >= 0.5) else "分歧⚠️"
            mean_s = round(hs.mean_fwd*100, 1) if not math.isnan(hs.mean_fwd) else "—"
            med_s = round(hs.median_fwd*100, 1) if not math.isnan(hs.median_fwd) else "—"
            tablerow(r, [f"{h} 天", mean_s, med_s, round(hs.win_hi*100, 1), knn_s, agree]); r += 1
        merged(r, f"      (kNN: 取特征最相似的 {st.knn_n} 个历史日, 看它们之后的胜率; 与分桶法'印证'则结论更可信)",
               font=Font(italic=True, size=9, color="777777"), height=16); r += 1

        # ④ 仓位建议拆解
        merged(r, "④  仓位建议 · 马丁回撤档位 × 统计置信度", fill=SEC_FILL,
               font=Font(bold=True, size=11, color="1F3D14"), height=20, border=True); r += 1
        amt_txt = ""
        if reserve_capital:
            amt_txt = f"   (单看 ≈ {reserve_capital*d.add_pct:,.0f}, 最终以组合分配页为准)"
        merged(r, f"      底仓档位 {d.tier_name} = {d.tier_base_pct*100:.0f}% 弹药   ×   "
                  f"统计置信乘数 {d.confidence:.2f}   =   建议加仓 {d.add_pct*100:.1f}% 弹药{amt_txt}",
               font=Font(size=10, bold=True), height=20, border=True); r += 1

        # ⑤ 全部决策依据 (逐条, 不省略)
        merged(r, "⑤  决策依据 (全部, 不省略)", fill=SEC_FILL,
               font=Font(bold=True, size=11, color="1F3D14"), height=20, border=True); r += 1
        for reason in d.reasons:
            merged(r, f"      ▪ {reason}", font=Font(size=10), height=18); r += 1

        ws.row_dimensions[r].height = 8
        r += 1

    # ====================== Sheet 2: 组合弹药分配 ====================== #
    allocations, dry = allocate_portfolio(decisions, reserve_capital)
    ws2 = wb.create_sheet("💰 组合弹药分配")
    ws2.sheet_view.showGridLines = False
    for i, w in enumerate([20, 10, 12, 14, 16, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.merge_cells("A1:F1")
    t = ws2["A1"]
    t.value = "💰 组合弹药分配 (按信心分配总加仓弹药 · 单资产封顶 35% · 信心不足留干火药)"
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=TITLE_FILL)
    t.alignment = CENTER
    ws2.row_dimensions[1].height = 28

    hdrs = ["资产", "市场", "决策", "占总弹药%", "建议金额", "原始意向(add_pct)%"]
    for col, h in enumerate(hdrs, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor=HDR_FILL)
        c.alignment = CENTER
        c.border = BORDER
    rr = 3
    for a in allocations:
        amount = f"{a['amount']:,.0f}" if a["amount"] is not None else "—"
        vals = [a["name"], a["market"], a["action"], round(a["share"]*100, 1),
                amount, round(a["add_pct"]*100, 1)]
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row=rr, column=col, value=v)
            c.border = BORDER
            c.alignment = CENTER if col > 1 else LEFT
            c.font = Font(size=10)
        ws2.cell(row=rr, column=3).fill = PatternFill("solid", fgColor=ACT_FILL.get(a["action"], "FFFFFF"))
        rr += 1
    # 干火药行
    dry_amt = f"{reserve_capital*dry:,.0f}" if reserve_capital else "—"
    dvals = ["🧱 干火药(留存观望)", "", "", round(dry*100, 1), dry_amt, ""]
    for col, v in enumerate(dvals, 1):
        c = ws2.cell(row=rr, column=col, value=v)
        c.border = BORDER
        c.alignment = CENTER if col > 1 else LEFT
        c.font = Font(size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="FFF2CC")
    rr += 2
    ws2.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
    note = ws2.cell(row=rr, column=1,
                    value="说明: add_pct=马丁档位×统计置信, 解释为'投占总弹药的比例'; 总和>100%则按信心归一到满仓, "
                          "≤100%则原样部署、其余留干火药; 任一资产封顶35%。决策仅供参考。")
    note.font = Font(italic=True, size=9, color="777777")
    note.alignment = LEFT


def _write_excel(decisions: List[AddDecision], reserve_capital: Optional[float] = None) -> None:
    """独立模式 (--excel): 自建 Workbook, 写入完整决策页, 落桌面并弹出。缺 openpyxl 自动跳过。"""
    try:
        from openpyxl import Workbook
    except Exception:
        print("   ⚠️ 未装 openpyxl, 跳过 Excel (pip install openpyxl)。")
        return
    wb = Workbook()
    default = wb.active
    build_midterm_sheets(wb, decisions, reserve_capital)
    wb.remove(default)  # 删掉自带的空白 Sheet
    path = os.path.expanduser("~/Desktop/中长线加仓决策看板.xlsx")
    wb.save(path)
    print(f"\n✅ Excel 看板已生成: {path}")
    try:
        if sys.platform == "darwin":
            os.system(f'open "{path}"')
    except Exception:
        pass


def main() -> None:
    args = sys.argv[1:]
    want_excel = "--excel" in args
    reserve_capital: Optional[float] = None
    if "--reserve" in args:
        i = args.index("--reserve")
        if i + 1 < len(args):
            try:
                reserve_capital = float(args[i + 1])
            except ValueError:
                print(f"⚠️ --reserve 后须跟数字, 收到 '{args[i + 1]}', 忽略。")
    if "--backtest" in args:
        i = args.index("--backtest")
        query = args[i + 1] if i + 1 < len(args) else "比特币"
        matched = _match_asset(query)
        if not matched:
            print(f"未找到匹配 '{query}' 的资产。可选: {list(STOCK_ASSETS) + list(CRYPTO_ASSETS)}")
            return
        name, code, mkt, is_crypto = matched[0]
        print(f"拉取 {name} 长历史日K 做回测验证 ...")
        df = fetch_crypto_daily(code) if is_crypto else fetch_stock_daily(code)
        if df is None or len(df) < WARMUP_BARS:
            print("数据不足, 无法回测。")
            return
        backtest_dca_vs_smart(df, name)
        return

    if "--asset" in args:
        i = args.index("--asset")
        query = args[i + 1] if i + 1 < len(args) else ""
        matched = _match_asset(query)
        if not matched:
            print(f"未找到匹配 '{query}' 的资产。")
            return
        scan(matched, want_excel, reserve_capital)
        return

    scan(_all_assets(), want_excel, reserve_capital)


if __name__ == "__main__":
    main()
